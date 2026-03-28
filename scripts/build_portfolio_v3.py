"""
BIST Portfolio V3 — Excel Workbook Builder
Builds BIST_PORTFOLIO_V3_FINAL.xlsx with:
  - dim_Tickers (with IRR_USD via XIRR)
  - fact_Returns (acquisition-date cumulative returns)
  - fact_Financials (with FCF columns)
  - fact_Dividends
  - Portfolio_CAGR (weighted metrics)
  - DAX_Measures (Power BI formulas)
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import math

# Color Palette (McKinsey)
NAVY = "051C2C"
BLUE = "2251FF"
TEAL = "009CDC"
GREEN = "2D8C3C"
RED = "E4002B"
YELLOW = "FFD100"
DARK = "333333"
MID = "58595B"
LIGHT = "F2F2F2"
WHITE = "FFFFFF"

header_font = Font(name="Calibri", bold=True, size=11, color=WHITE)
title_font = Font(name="Calibri", bold=True, size=12, color=NAVY)
data_font = Font(name="Calibri", size=10, color=DARK)
green_font = Font(name="Calibri", size=10, color=GREEN, bold=True)
red_font = Font(name="Calibri", size=10, color=RED, bold=True)
muted_font = Font(name="Calibri", size=9, color=MID)

navy_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
light_fill = PatternFill(start_color=LIGHT, end_color=LIGHT, fill_type="solid")
yellow_fill = PatternFill(start_color="FFF8D6", end_color="FFF8D6", fill_type="solid")
green_fill = PatternFill(start_color="EBF5ED", end_color="EBF5ED", fill_type="solid")
red_fill = PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid")

thin_border = Border(bottom=Side(style="thin", color="DDDDDD"))


def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")


def auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 25)


tickers_data = [
    ("HALKB", "Halk Bankasi", "Bankacilik", 0.9149, "2017-02-03", 10.5, 2.82, 33.92, 5630, -0.133, -0.72),
    ("VAKBN", "Vakiflar Bankasi", "Bankacilik", 0.7326, "2020-05-15", 4.12, 0.60, 26.50, 5190, 0.005, 0.03),
    ("THYAO", "Turk Hava Yollari", "Ulastirma", 0.4912, "2017-02-03", 5.83, 1.56, 285.50, 4240, 0.175, 3.20),
    ("TTKOM", "Turk Telekomunikasyon", "Haberlesme", 0.5500, "2024-09-01", 52.80, 1.58, 58.30, 2580, -0.103, -0.13),
    ("TURSG", "Turkiye Sigorta", "Sigorta", 0.8100, "2024-06-15", 22.50, 0.69, 30.40, 2230, 0.022, 0.03),
    ("TRALT", "Turk Altin", "Madencilik", 0.4801, "2024-08-19", 31.20, 0.93, 48.10, 1470, 0.155, 0.22),
    ("TCELL", "Turkcell", "Haberlesme", 0.2620, "2024-03-10", 68.50, 2.13, 110.20, 1250, 0.105, 0.20),
    ("TRMET", "Turk Metal", "Madencilik", 0.5225, "2024-08-19", 18.40, 0.55, 25.90, 493, 0.082, 0.11),
    ("TRENJ", "Turkerler Enerji", "Madencilik", 0.6212, "2024-08-19", 5.90, 0.18, 10.50, 350, 0.288, 0.41),
    ("KRDMD", "Kardemir Demir Celik", "Ana Metal", 0.0441, "2022-12-05", 16.20, 0.87, 23.30, 30, -0.142, -0.37),
    ("KAYSE", "Kayseri Seker Fabrikasi", "Gida", 0.0941, "2023-05-12", 28.50, 1.44, 32.20, 29, -0.223, -0.49),
]

dividends_data = [
    ("TCELL", "2024-12-15", 2.5679, 34.80),
    ("TCELL", "2025-06-15", 1.5455, 38.50),
    ("TCELL", "2025-12-15", 1.5455, 42.00),
    ("THYAO", "2025-06-15", 2.9257, 38.50),
    ("THYAO", "2025-09-15", 2.9257, 40.20),
    ("KRDMD", "2023-06-15", 0.0987, 23.50),
    ("KRDMD", "2023-11-15", 0.0987, 28.80),
    ("TURSG", "2024-08-15", 0.18, 33.20),
    ("TURSG", "2025-08-15", 0.17, 40.50),
    ("KAYSE", "2023-09-15", 0.045, 26.90),
    ("KAYSE", "2023-11-15", 0.2141, 28.80),
]

returns_sample = [
    ("THYAO", "2017-Q1", "2017-03-31", 20171, 6.10, 1.62, 0.046, 0.038),
    ("THYAO", "2024-Q4", "2024-12-31", 20244, 262.00, 7.63, 43.93, 3.89),
    ("THYAO", "2025-Q4", "2025-12-31", 20254, 285.50, 6.65, 48.09, 3.26),
    ("HALKB", "2017-Q1", "2017-03-31", 20171, 11.20, 2.98, 0.067, 0.057),
    ("HALKB", "2024-Q4", "2024-12-31", 20244, 30.50, 0.89, 1.90, -0.68),
    ("HALKB", "2025-Q4", "2025-12-31", 20254, 33.92, 0.79, 2.23, -0.72),
]

financials_sample = [
    ("THYAO", "2025-Q4", "2025-12-31", 20254, 270000, 6280, 37800, 881, 40300, 935, 0.14),
    ("THYAO", "2025-Q3", "2025-09-30", 20253, 252000, 6100, 35200, 850, 38000, 920, 0.14),
    ("HALKB", "2025-Q4", "2025-12-31", 20254, 185000, 4300, 12500, 291, None, None, 0.068),
    ("VAKBN", "2025-Q4", "2025-12-31", 20254, 162000, 3770, 11800, 275, None, None, 0.073),
]


def xirr_approx(cashflows, dates):
    if not cashflows or len(cashflows) < 2:
        return None

    def npv(rate, cfs, ds):
        d0 = ds[0]
        total = 0
        for cf, d in zip(cfs, ds):
            years = (d - d0).days / 365.25
            if rate <= -1 and years != 0:
                return float("inf")
            total += cf / ((1 + rate) ** years)
        return total

    lo, hi = -0.5, 5.0
    for _ in range(200):
        mid = (lo + hi) / 2
        v = npv(mid, cashflows, dates)
        if abs(v) < 0.01:
            return mid
        if v > 0:
            lo = mid
        else:
            hi = mid
    return (lo + hi) / 2


def compute_irr_for_ticker(ticker, acq_date_str, acq_price_usd, current_price_usd):
    acq_date = datetime.strptime(acq_date_str, "%Y-%m-%d").date()
    cashflows = [-acq_price_usd]
    dates = [acq_date]

    for div_ticker, div_date_str, dps_try, usdtry in dividends_data:
        if div_ticker == ticker:
            div_date = datetime.strptime(div_date_str, "%Y-%m-%d").date()
            if div_date > acq_date:
                cashflows.append(dps_try / usdtry)
                dates.append(div_date)

    end_date = date(2025, 12, 31)
    cashflows.append(current_price_usd)
    dates.append(end_date)

    return xirr_approx(cashflows, dates)


USDTRY_CURRENT = 42.9557

wb = openpyxl.Workbook()

# Sheet 1: dim_Tickers
ws = wb.active
ws.title = "dim_Tickers"
headers = [
    "Ticker", "Company_Name", "Sector", "Ownership_Pct",
    "Acq_Date", "Acq_Price_TRY", "Acq_Price_USD",
    "Current_Price_TRY", "Your_Value_USD_M",
    "CAGR_USD", "Total_Return_USD", "IRR_USD"
]
ws.append(headers)
style_header_row(ws, 1, len(headers))

for row in tickers_data:
    ticker, company, sector, own, acq_date, acq_try, acq_usd, cur_try, val_usd, cagr, ret = row
    cur_usd = cur_try / USDTRY_CURRENT
    irr = compute_irr_for_ticker(ticker, acq_date, acq_usd, cur_usd)
    ws.append([
        ticker, company, sector, own,
        acq_date, acq_try, acq_usd,
        cur_try, val_usd,
        round(cagr, 4), round(ret, 4),
        round(irr, 4) if irr else None
    ])
    r = ws.max_row
    for c in [10, 11, 12]:
        cell = ws.cell(row=r, column=c)
        if cell.value and cell.value > 0:
            cell.font = green_font
        elif cell.value and cell.value < 0:
            cell.font = red_font
    ws.cell(row=r, column=4).number_format = "0.00%"
    for c in [10, 11, 12]:
        ws.cell(row=r, column=c).number_format = "0.00%"
auto_width(ws)

# Sheet 2: fact_Returns
ws2 = wb.create_sheet("fact_Returns")
ret_headers = ["Ticker", "Q_Label", "Q_Date", "Sort_Order", "Price_TRY", "Price_USD", "Cumul_Nom_TRY", "Cumul_Nom_USD"]
ws2.append(ret_headers)
style_header_row(ws2, 1, len(ret_headers))
for row in returns_sample:
    ws2.append(list(row))
auto_width(ws2)

# Sheet 3: fact_Financials
ws3 = wb.create_sheet("fact_Financials")
fin_headers = [
    "Ticker", "Q_Label", "Q_Date", "Sort_Order",
    "Revenue_TRY_M", "Revenue_USD_M", "Net_Income_TRY_M", "Net_Income_USD_M",
    "EBITDA_TRY_M", "EBITDA_USD_M", "Net_Margin",
    "Op_CF_TRY_M", "Op_CF_USD_M", "Capex_TRY_M", "Capex_USD_M",
    "FCF_TRY_M", "FCF_USD_M"
]
ws3.append(fin_headers)
style_header_row(ws3, 1, len(fin_headers))
for row in financials_sample:
    data = list(row) + [None, None, None, None, None, None]
    ws3.append(data)
auto_width(ws3)

# Sheet 4: fact_Dividends
ws4 = wb.create_sheet("fact_Dividends")
div_headers = ["Ticker", "Payment_Date", "Net_DPS_TRY", "USDTRY_Rate", "Net_DPS_USD", "Total_Cash_Div_TRY_M", "Your_Div_Income_TRY_M"]
ws4.append(div_headers)
style_header_row(ws4, 1, len(div_headers))
for div_ticker, div_date, dps_try, usdtry in dividends_data:
    dps_usd = round(dps_try / usdtry, 6)
    ws4.append([div_ticker, div_date, dps_try, usdtry, dps_usd, None, None])
auto_width(ws4)

# Sheet 5: Portfolio_CAGR
ws5 = wb.create_sheet("Portfolio_CAGR")
cagr_headers = ["Ticker", "Weight", "CAGR_USD", "IRR_USD", "Total_Return_USD", "CAGR_Contribution", "IRR_Contribution"]
ws5.append(cagr_headers)
style_header_row(ws5, 1, len(cagr_headers))
total_value = sum(r[8] for r in tickers_data)
for row in tickers_data:
    ticker = row[0]
    val = row[8]
    cagr = row[9]
    ret = row[10]
    weight = val / total_value
    cur_usd = row[7] / USDTRY_CURRENT
    irr = compute_irr_for_ticker(ticker, row[4], row[6], cur_usd)
    ws5.append([
        ticker, round(weight, 4), round(cagr, 4),
        round(irr, 4) if irr else None, round(ret, 4),
        round(weight * cagr, 5),
        round(weight * irr, 5) if irr else None,
    ])
weighted_cagr = sum((r[8] / total_value) * r[9] for r in tickers_data)
ws5.append([])
ws5.append(["PORTFOLIO", 1.0, round(weighted_cagr, 4), None, None, round(weighted_cagr, 4), None])
r = ws5.max_row
for c in range(1, 8):
    ws5.cell(row=r, column=c).font = Font(name="Calibri", bold=True, size=11, color=NAVY)
    ws5.cell(row=r, column=c).fill = yellow_fill
auto_width(ws5)

# Sheet 6: DAX_Measures
ws6 = wb.create_sheet("DAX_Measures")
ws6.append(["Measure_Name", "DAX_Formula", "Description"])
style_header_row(ws6, 1, 3)
dax_measures = [
    ("Weighted_CAGR", 'SUMX(dim_Tickers, dim_Tickers[CAGR_USD] * dim_Tickers[Your_Value_USD_M]) / SUM(dim_Tickers[Your_Value_USD_M])', "Value-weighted portfolio CAGR"),
    ("Weighted_IRR", 'SUMX(dim_Tickers, dim_Tickers[IRR_USD] * dim_Tickers[Your_Value_USD_M]) / SUM(dim_Tickers[Your_Value_USD_M])', "Value-weighted portfolio IRR"),
    ("Weighted_Return", 'SUMX(dim_Tickers, dim_Tickers[Total_Return_USD] * dim_Tickers[Your_Value_USD_M]) / SUM(dim_Tickers[Your_Value_USD_M])', "Value-weighted total return"),
    ("Portfolio_NAV", 'SUM(dim_Tickers[Your_Value_USD_M])', "Total portfolio NAV in USD millions"),
    ("Latest_Revenue_USD", 'CALCULATE(SUM(fact_Financials[Revenue_USD_M]), FILTER(ALL(fact_Financials), fact_Financials[Q_Date] = MAX(fact_Financials[Q_Date])))', "Latest quarter revenue"),
    ("Latest_EBITDA_USD", 'CALCULATE(SUM(fact_Financials[EBITDA_USD_M]), FILTER(ALL(fact_Financials), fact_Financials[Q_Date] = MAX(fact_Financials[Q_Date])))', "Latest quarter EBITDA"),
    ("Latest_FCF_USD", 'CALCULATE(SUM(fact_Financials[FCF_USD_M]), FILTER(ALL(fact_Financials), fact_Financials[Q_Date] = MAX(fact_Financials[Q_Date])))', "Latest quarter FCF"),
    ("Latest_Net_Margin", 'CALCULATE(AVERAGE(fact_Financials[Net_Margin]), FILTER(ALL(fact_Financials), fact_Financials[Q_Date] = MAX(fact_Financials[Q_Date])))', "Latest quarter avg net margin"),
    ("THYAO_Return", 'CALCULATE(MAX(dim_Tickers[Total_Return_USD]), dim_Tickers[Ticker] = "THYAO", ALL(dim_Tickers))', "THYAO total USD return"),
    ("HALKB_Return", 'CALCULATE(MAX(dim_Tickers[Total_Return_USD]), dim_Tickers[Ticker] = "HALKB", ALL(dim_Tickers))', "HALKB total USD return"),
    ("Your_Dividend_Income", 'SUMX(fact_Dividends, fact_Dividends[Your_Div_Income_TRY_M])', "Total ownership-weighted dividend income"),
    ("TRY_Discount_Rate_Value", "SELECTEDVALUE('TRY Discount Rate'[TRY Discount Rate], 0.45)", "What-If param: TRY discount rate"),
    ("USD_Discount_Rate_Value", "SELECTEDVALUE('USD Discount Rate'[USD Discount Rate], 0.055)", "What-If param: USD discount rate"),
    ("Quarterly_Rate_TRY", "POWER(1 + [TRY_Discount_Rate_Value], 1/4) - 1", "Quarterly compound TRY rate"),
    ("Quarterly_Rate_USD", "POWER(1 + [USD_Discount_Rate_Value], 1/4) - 1", "Quarterly compound USD rate"),
    ("Revenue_Real", "DIVIDE(SUM(fact_Financials[Revenue_USD_M]), POWER(1 + [Quarterly_Rate_USD], MAX(fact_Financials[Sort_Order]) - MIN(fact_Financials[Sort_Order])))", "Inflation-adjusted revenue"),
]
for name, formula, desc in dax_measures:
    ws6.append([name, formula, desc])
auto_width(ws6)

# Sheet 7: param_Rates
ws7 = wb.create_sheet("param_Rates")
ws7.append(["Rate_Type", "Currency", "Annual_Rate_Pct", "Quarterly_Rate"])
style_header_row(ws7, 1, 4)
ws7.append(["Discount Rate", "TRY", 0.45, round((1.45 ** 0.25) - 1, 6)])
ws7.append(["Discount Rate", "USD", 0.055, round((1.055 ** 0.25) - 1, 6)])
auto_width(ws7)

output_path = "scripts/BIST_PORTFOLIO_V3_FINAL.xlsx"
wb.save(output_path)
print(f"Workbook saved to {output_path}")
print(f"Sheets: {wb.sheetnames}")
print(f"dim_Tickers: {ws.max_row - 1} rows")
print(f"Portfolio CAGR (weighted): {weighted_cagr:.4f} = {weighted_cagr*100:.2f}%")
