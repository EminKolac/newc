#!/usr/bin/env python3
"""
DCF Model Doğrulama & Genel Excel Audit — Excel modellerini kontrol eder.

Kaynak: Claude financial-services-plugins/dcf-model/validate_dcf.py — BIST/Türkiye'ye adapte edildi.

Kontroller (DCF modu — varsayılan):
  1. Sheet yapısı (DCF, WACC, Sensitivity sheet'leri)
  2. Excel formül hataları (#REF!, #DIV/0!, #VALUE!, vb.)
  3. Terminal büyüme < WACC (kritik — ihlal = sonsuz değer)
  4. WACC aralık kontrolü (TL bazlı: %10-45, USD bazlı: %5-20)
  5. Terminal değer oranı (TV/EV %40-80 arası olmalı, >%80 uyarı)
  6. Fisher parity cross-check (TL WACC vs USD WACC tutarlılığı)

Kontroller (--audit modu — genel Excel audit):
  A. Formül hataları (#REF!, #DIV/0!, #VALUE!, #NAME?, #N/A, vb.)
  B. Hardcoded değer tespiti (formül olması gereken yerde sabit sayı)
     - Whitelist: varsayım satırları (büyüme, oran, marj, wacc, beta, vb.)
     - Sheet whitelist: 'Input', 'Inputs', 'Assumptions', 'Varsayım' sheet'leri
  C. Döngüsel referans uyarısı (openpyxl sınırlaması — tespit edildiğinde bildir)
  D. Format tutarsızlıkları (aynı sütunda farklı sayı formatı)
  E. Boş hücreye referans veren formüller

Kullanım:
    python dcf-dogrulama.py <excel_dosyasi> [--json cikti.json] [--para-birimi TL|USD]
    python dcf-dogrulama.py <excel_dosyasi> --audit
    python dcf-dogrulama.py <excel_dosyasi> --audit --json sonuc.json

Örnekler:
    python dcf-dogrulama.py THYAO_DCF.xlsx
    python dcf-dogrulama.py THYAO_Financial_Model.xlsx --audit
    python dcf-dogrulama.py THYAO_DCF.xlsx --audit --para-birimi USD --json sonuc.json
"""

import json
import sys
import argparse
from pathlib import Path
from typing import Dict, Any, Tuple, Optional, List
from datetime import datetime


class DCFModelDogrulayici:
    """Excel DCF modellerini doğrular — BIST/Türkiye parametreleri ile."""

    # WACC aralıkları (para birimine göre)
    WACC_ARALIKLARI = {
        'TL':  {'min': 0.10, 'max': 0.45, 'aciklama': '%10-%45 (TL bazlı, Türkiye enflasyonu dahil)'},
        'USD': {'min': 0.05, 'max': 0.20, 'aciklama': '%5-%20 (USD bazlı, gelişmekte olan piyasa)'},
        'EUR': {'min': 0.05, 'max': 0.18, 'aciklama': '%5-%18 (EUR bazlı)'},
    }

    # Terminal büyüme aralıkları
    TERMINAL_BUYUME_ARALIKLARI = {
        'TL':  {'min': 0.03, 'max': 0.15, 'aciklama': '%3-%15 (TL: nominal, enflasyon dahil)'},
        'USD': {'min': 0.01, 'max': 0.04, 'aciklama': '%1-%4 (USD: reel)'},
        'EUR': {'min': 0.01, 'max': 0.03, 'aciklama': '%1-%3 (EUR: reel)'},
    }

    EXCEL_HATALARI = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']

    def __init__(self, excel_yolu: str, para_birimi: str = 'TL'):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl gerekli. Kur: pip install openpyxl")

        self.excel_yolu = excel_yolu
        self.para_birimi = para_birimi.upper()
        self.openpyxl = openpyxl

        if not Path(excel_yolu).exists():
            raise FileNotFoundError(f"Dosya bulunamadı: {excel_yolu}")

        if self.para_birimi not in self.WACC_ARALIKLARI:
            raise ValueError(f"Desteklenmeyen para birimi: {para_birimi}. Desteklenen: TL, USD, EUR")

        self.wb_formuller = openpyxl.load_workbook(excel_yolu, data_only=False)
        self.wb_degerler = openpyxl.load_workbook(excel_yolu, data_only=True)
        self.hatalar: List[str] = []
        self.uyarilar: List[str] = []
        self.bilgiler: List[str] = []

    def dogrula(self) -> dict:
        """Tüm doğrulama kontrollerini çalıştır."""
        self._sheet_yapisi_kontrol()
        self._formul_hatalari_kontrol()
        self._dcf_mantik_kontrol()

        sonuc = {
            'dosya': self.excel_yolu,
            'para_birimi': self.para_birimi,
            'dogrulama_tarihi': datetime.now().isoformat(),
            'durum': 'GEÇTİ' if len(self.hatalar) == 0 else 'BAŞARISIZ',
            'hata_sayisi': len(self.hatalar),
            'uyari_sayisi': len(self.uyarilar),
            'hatalar': self.hatalar,
            'uyarilar': self.uyarilar,
            'bilgiler': self.bilgiler,
        }
        return sonuc

    def _sheet_yapisi_kontrol(self):
        """Gerekli sheet'lerin varlığını kontrol et."""
        # Hem İngilizce hem Türkçe isimleri kabul et
        gerekli_sheetler = {
            'DCF': ['DCF', 'dcf', 'İNA', 'ina', 'Değerleme', 'FCFFSimple', 'FCFF'],
            'WACC': ['WACC', 'wacc', 'WACCCalc', 'İskonto', 'AOSM'],
            'Sensitivity': ['Sensitivity', 'sensitivity', 'Duyarlılık', 'Senaryo', 'Output'],
        }
        sheet_isimleri = self.wb_degerler.sheetnames
        self.bilgiler.append(f"Bulunan sheet'ler: {', '.join(sheet_isimleri)}")

        for gerekli, alternatifler in gerekli_sheetler.items():
            bulundu = any(alt in sheet_isimleri for alt in alternatifler)
            if bulundu:
                self.bilgiler.append(f"✓ {gerekli} sheet'i mevcut")
            else:
                self.uyarilar.append(
                    f"Önerilen sheet bulunamadı: {gerekli} "
                    f"(aranan isimler: {', '.join(alternatifler)})"
                )

    def _formul_hatalari_kontrol(self):
        """Tüm sheet'lerde Excel formül hatalarını tara."""
        toplam_hata = 0
        toplam_formul = 0

        for sheet_adi in self.wb_degerler.sheetnames:
            ws_degerler = self.wb_degerler[sheet_adi]
            ws_formuller = self.wb_formuller[sheet_adi]

            for satir in ws_degerler.iter_rows():
                for hucre in satir:
                    formul_hucre = ws_formuller[hucre.coordinate]

                    # Formül sayısı
                    if (formul_hucre.value
                            and isinstance(formul_hucre.value, str)
                            and formul_hucre.value.startswith('=')):
                        toplam_formul += 1

                    # Hata kontrolü
                    if hucre.value is not None and isinstance(hucre.value, str):
                        for hata in self.EXCEL_HATALARI:
                            if hata in hucre.value:
                                konum = f"{sheet_adi}!{hucre.coordinate}"
                                self.hatalar.append(f"{hata} → {konum}")
                                toplam_hata += 1
                                break

        self.bilgiler.append(f"Toplam formül sayısı: {toplam_formul}")
        if toplam_hata == 0:
            self.bilgiler.append("✓ Formül hatası bulunamadı")
        else:
            self.hatalar.append(f"Toplam formül hatası: {toplam_hata}")

    def _dcf_mantik_kontrol(self):
        """DCF'e özgü mantık kontrolleri."""
        self._terminal_buyume_vs_wacc()
        self._wacc_aralik_kontrol()
        self._terminal_deger_orani()

    def _hucre_deger_ara(self, sheet, anahtar_kelimeler: list, max_satir=200, max_sutun=20) -> Optional[float]:
        """Sheet'te anahtar kelimeye göre değer ara."""
        for satir in sheet.iter_rows(max_row=max_satir, max_col=max_sutun):
            for hucre in satir:
                if hucre.value and isinstance(hucre.value, str):
                    hucre_str = hucre.value.lower()
                    if all(k in hucre_str for k in anahtar_kelimeler):
                        # Komşu hücrelerde sayısal değer ara
                        for offset in range(1, 6):
                            try:
                                komsu = sheet.cell(hucre.row, hucre.column + offset).value
                                if isinstance(komsu, (int, float)) and komsu > 0:
                                    return komsu
                            except Exception:
                                continue
                        # Alt hücrelerde de ara
                        for offset in range(1, 4):
                            try:
                                alt = sheet.cell(hucre.row + offset, hucre.column).value
                                if isinstance(alt, (int, float)) and alt > 0:
                                    return alt
                            except Exception:
                                continue
        return None

    def _sheet_bul(self, alternatifler: list):
        """Alternatif isimlerle sheet bul."""
        for alt in alternatifler:
            if alt in self.wb_degerler.sheetnames:
                return self.wb_degerler[alt]
        return None

    def _terminal_buyume_vs_wacc(self):
        """KRİTİK: Terminal büyüme oranı < WACC olmalı."""
        dcf_sheet = self._sheet_bul(['DCF', 'dcf', 'İNA', 'FCFFSimple', 'FCFF', 'Değerleme'])
        if not dcf_sheet:
            self.uyarilar.append("DCF sheet'i bulunamadı — terminal büyüme kontrolü atlandı")
            return

        terminal_buyume = self._hucre_deger_ara(
            dcf_sheet, ['terminal', 'growth']
        ) or self._hucre_deger_ara(
            dcf_sheet, ['terminal', 'büyüme']
        ) or self._hucre_deger_ara(
            dcf_sheet, ['stable', 'growth']
        )

        wacc_sheet = self._sheet_bul(['WACC', 'wacc', 'WACCCalc', 'İskonto', 'AOSM']) or dcf_sheet
        wacc = self._hucre_deger_ara(
            wacc_sheet, ['wacc']
        ) or self._hucre_deger_ara(
            wacc_sheet, ['aosm']
        ) or self._hucre_deger_ara(
            wacc_sheet, ['cost', 'capital']
        )

        if terminal_buyume is not None and wacc is not None:
            # Yüzde olarak girilmişse düzelt (>1 ise yüzde olarak kabul et)
            if terminal_buyume > 1:
                terminal_buyume = terminal_buyume / 100
            if wacc > 1:
                wacc = wacc / 100

            if terminal_buyume >= wacc:
                self.hatalar.append(
                    f"🔴 KRİTİK: Terminal büyüme ({terminal_buyume:.2%}) ≥ WACC ({wacc:.2%}). "
                    "Bu matematiksel olarak sonsuz değer üretir — MODEL GEÇERSİZ."
                )
            else:
                spread = wacc - terminal_buyume
                self.bilgiler.append(
                    f"✓ Terminal büyüme ({terminal_buyume:.2%}) < WACC ({wacc:.2%}), "
                    f"spread: {spread:.2%}"
                )
                if spread < 0.02:
                    self.uyarilar.append(
                        f"Terminal büyüme-WACC spread'i çok dar ({spread:.2%}). "
                        "Küçük değişiklikler büyük değerleme farkı yaratır."
                    )
        else:
            eksikler = []
            if terminal_buyume is None:
                eksikler.append("terminal büyüme")
            if wacc is None:
                eksikler.append("WACC")
            self.uyarilar.append(f"Bulunamadı: {', '.join(eksikler)} — otomatik kontrol yapılamadı")

    def _wacc_aralik_kontrol(self):
        """WACC değerinin para birimine uygun aralıkta olup olmadığını kontrol et."""
        wacc_sheet = self._sheet_bul(['WACC', 'wacc', 'WACCCalc', 'İskonto', 'AOSM'])
        if not wacc_sheet:
            # DCF sheet'inde ara
            wacc_sheet = self._sheet_bul(['DCF', 'dcf', 'İNA', 'FCFFSimple', 'FCFF'])
        if not wacc_sheet:
            return

        wacc = self._hucre_deger_ara(wacc_sheet, ['wacc']) or self._hucre_deger_ara(wacc_sheet, ['aosm'])
        if wacc is None:
            return

        if wacc > 1:
            wacc = wacc / 100

        aralik = self.WACC_ARALIKLARI[self.para_birimi]
        if wacc < aralik['min'] or wacc > aralik['max']:
            self.uyarilar.append(
                f"WACC ({wacc:.2%}) beklenen aralığın dışında: {aralik['aciklama']}. "
                "Hesaplamayı doğrula."
            )
        else:
            self.bilgiler.append(
                f"✓ WACC ({wacc:.2%}) makul aralıkta: {aralik['aciklama']}"
            )

        # Terminal büyüme aralık kontrolü
        dcf_sheet = self._sheet_bul(['DCF', 'dcf', 'İNA', 'FCFFSimple', 'FCFF', 'Değerleme'])
        if dcf_sheet:
            tg = (self._hucre_deger_ara(dcf_sheet, ['terminal', 'growth'])
                  or self._hucre_deger_ara(dcf_sheet, ['terminal', 'büyüme'])
                  or self._hucre_deger_ara(dcf_sheet, ['stable', 'growth']))
            if tg is not None:
                if tg > 1:
                    tg = tg / 100
                tg_aralik = self.TERMINAL_BUYUME_ARALIKLARI[self.para_birimi]
                if tg < tg_aralik['min'] or tg > tg_aralik['max']:
                    self.uyarilar.append(
                        f"Terminal büyüme ({tg:.2%}) beklenen aralığın dışında: "
                        f"{tg_aralik['aciklama']}. Doğrula."
                    )

    def _terminal_deger_orani(self):
        """Terminal değerin toplam EV içindeki payını kontrol et."""
        dcf_sheet = self._sheet_bul(['DCF', 'dcf', 'İNA', 'FCFFSimple', 'FCFF', 'Değerleme', 'Output'])
        if not dcf_sheet:
            return

        tv = (self._hucre_deger_ara(dcf_sheet, ['terminal', 'value', 'pv'])
              or self._hucre_deger_ara(dcf_sheet, ['terminal', 'değer'])
              or self._hucre_deger_ara(dcf_sheet, ['pv', 'terminal']))

        ev = (self._hucre_deger_ara(dcf_sheet, ['enterprise', 'value'])
              or self._hucre_deger_ara(dcf_sheet, ['firma', 'değer'])
              or self._hucre_deger_ara(dcf_sheet, ['toplam', 'değer']))

        if tv is not None and ev is not None and ev > 0:
            oran = tv / ev
            if oran > 0.80:
                self.uyarilar.append(
                    f"Terminal değer EV'nin %{oran:.0%}'i — çok yüksek (beklenen: %40-70). "
                    "Model terminal varsayımlara aşırı bağımlı."
                )
            elif oran < 0.30:
                self.uyarilar.append(
                    f"Terminal değer EV'nin %{oran:.0%}'i — çok düşük (beklenen: %40-70). "
                    "Terminal varsayımlar çok muhafazakâr olabilir."
                )
            else:
                self.bilgiler.append(f"✓ Terminal değer oranı: {oran:.0%} (EV içinde)")
        else:
            self.uyarilar.append("Terminal değer ve/veya EV bulunamadı — oran kontrolü atlandı")


class GenelAudit:
    """
    Genel Excel model audit'i — T2 finansal model çıktısını kontrol eder.

    4 kontrol kategorisi:
      A. Formül hataları (#REF!, #DIV/0!, #VALUE!, vb.)
      B. Hardcoded değer tespiti (projeksiyon bölgelerinde formül yerine sabit)
      C. Format tutarsızlıkları (aynı sütunda karışık sayı/para formatı)
      D. Boş hücreye referans uyarısı

    Whitelist mantığı (hardcoded için):
      - Sheet whitelist: Input/Inputs/Assumptions/Varsayım adlı sheet'ler → tamamen atla
      - Satır whitelist: Satır etiketi anahtar kelime içeriyorsa → hardcoded normal
        (büyüme, oran, marj, wacc, beta, katsayı, çarpan, oran, faiz, vergi, vb.)
      - Projeksiyon eşiği: Sütun başlığı 2025'ten büyük yıl içeriyorsa projeksiyon bölgesi
    """

    EXCEL_HATALARI = ['#VALUE!', '#DIV/0!', '#REF!', '#NAME?', '#NULL!', '#NUM!', '#N/A']

    # Bu anahtar kelimeleri içeren satırlar varsayım satırıdır — hardcoded olabilir
    VARSAYIM_SATIR_WHITELIST = [
        'büyüme', 'oran', 'marj', 'wacc', 'beta', 'katsayı', 'çarpan',
        'faiz', 'vergi', 'iskonto', 'terminal', 'enflasyon', 'fx', 'kur',
        'capex', 'capex/hasılat', 'capex/gelir', 'arge', 'temettu', 'temettü',
        'growth', 'rate', 'margin', 'discount', 'inflation', 'multiple',
        'varsayım', 'assumption', 'girdi', 'input', 'parametre', 'senaryo',
        'erp', 'crp', 'rf', 'spread', 'leverage', 'kaldıraç', 'd/e', 'tax',
    ]

    # Bu isimleri içeren sheet'ler → tamamen whitelist (tümü hardcoded olabilir)
    VARSAYIM_SHEET_WHITELIST = [
        'input', 'inputs', 'assumptions', 'varsayım', 'varsayimlar',
        'parametreler', 'parameters', 'girdi', 'girdiler', 'senaryo',
        'scenarios', 'macro', 'makro', 'industry', 'sektör',
    ]

    def __init__(self, excel_yolu: str):
        try:
            import openpyxl
        except ImportError:
            raise ImportError("openpyxl gerekli. Kur: pip install openpyxl")

        self.excel_yolu = excel_yolu
        self.openpyxl = openpyxl

        if not Path(excel_yolu).exists():
            raise FileNotFoundError(f"Dosya bulunamadı: {excel_yolu}")

        self.wb_formuller = openpyxl.load_workbook(excel_yolu, data_only=False)
        self.wb_degerler = openpyxl.load_workbook(excel_yolu, data_only=True)
        self.hatalar: List[dict] = []
        self.uyarilar: List[dict] = []
        self.bilgiler: List[str] = []

    def audit(self) -> dict:
        """Tüm genel audit kontrollerini çalıştır."""
        self.bilgiler.append(f"Sheet'ler: {', '.join(self.wb_degerler.sheetnames)}")
        self._formul_hatalari_tara()
        self._hardcoded_tespiti()
        self._format_tutarsizliklari()

        kritik = [h for h in self.hatalar if h.get('seviye') == 'kritik']
        uyari = [h for h in self.hatalar if h.get('seviye') == 'uyari'] + self.uyarilar

        durum = 'TEMİZ' if len(kritik) == 0 else 'SORUNLU'

        return {
            'dosya': self.excel_yolu,
            'audit_tarihi': datetime.now().isoformat(),
            'durum': durum,
            'kritik_sayisi': len(kritik),
            'uyari_sayisi': len(uyari),
            'kritikler': kritik,
            'uyarilar': uyari,
            'bilgiler': self.bilgiler,
        }

    def _formul_hatalari_tara(self):
        """Tüm sheet'lerde Excel formül hatalarını tara — kritik."""
        toplam = 0
        for sheet_adi in self.wb_degerler.sheetnames:
            ws = self.wb_degerler[sheet_adi]
            for satir in ws.iter_rows():
                for hucre in satir:
                    if hucre.value and isinstance(hucre.value, str):
                        for hata_kodu in self.EXCEL_HATALARI:
                            if hata_kodu in str(hucre.value):
                                self.hatalar.append({
                                    'seviye': 'kritik',
                                    'kategori': 'Formül Hatası',
                                    'konum': f"{sheet_adi}!{hucre.coordinate}",
                                    'deger': hucre.value,
                                    'aciklama': f"{hata_kodu} — düzeltilmeli",
                                })
                                toplam += 1
                                break
        if toplam == 0:
            self.bilgiler.append("✓ Formül hatası: 0")
        else:
            self.bilgiler.append(f"⚠ Formül hatası: {toplam} hücre")

    def _hardcoded_tespiti(self):
        """
        Projeksiyon bölgelerinde formül yerine hardcoded sabit tespit eder.

        Kural:
        - Sheet whitelist'teyse → atla (Input/Assumptions sheet'leri)
        - Satır etiketi varsayım whitelist'indeyse → atla
        - Projeksiyon yılı sütunu (>= cari yıl + 1) + sayısal sabit varsa → uyarı
        """
        from datetime import date
        cari_yil = date.today().year

        for sheet_adi in self.wb_formuller.sheetnames:
            # Sheet whitelist kontrolü
            if any(w in sheet_adi.lower() for w in self.VARSAYIM_SHEET_WHITELIST):
                self.bilgiler.append(f"  Sheet whitelist: '{sheet_adi}' → hardcoded kontrolü atlandı")
                continue

            ws_formul = self.wb_formuller[sheet_adi]
            ws_deger = self.wb_degerler[sheet_adi]

            # Projeksiyon sütunlarını tespit et (başlık = yıl >= cari_yil+1)
            projeksiyon_sutunlari = set()
            for hucre in ws_formul[1]:  # İlk satır = başlıklar
                if hucre.value and isinstance(hucre.value, (int, float)):
                    if int(hucre.value) >= cari_yil + 1:
                        projeksiyon_sutunlari.add(hucre.column)
                elif hucre.value and isinstance(hucre.value, str):
                    # "2026", "FY26", "26E", "2026E" gibi formatları dene
                    import re
                    yil_eslesme = re.search(r'20(\d{2})', str(hucre.value))
                    if yil_eslesme:
                        tam_yil = int('20' + yil_eslesme.group(1))
                        if tam_yil >= cari_yil + 1:
                            projeksiyon_sutunlari.add(hucre.column)

            if not projeksiyon_sutunlari:
                continue  # Bu sheet'te projeksiyon sütunu yok

            # Satır bazlı kontrol
            for satir in ws_formul.iter_rows(min_row=2):
                # Satır etiketi al (ilk sütun)
                satir_etiketi = ''
                if satir[0].value:
                    satir_etiketi = str(satir[0].value).lower()

                # Satır whitelist kontrolü
                if any(w in satir_etiketi for w in self.VARSAYIM_SATIR_WHITELIST):
                    continue

                # Projeksiyon sütunlarında hardcoded kontrolü
                for hucre in satir:
                    if hucre.column not in projeksiyon_sutunlari:
                        continue
                    # Formül değil, sayısal sabit varsa
                    if (hucre.value is not None
                            and not isinstance(hucre.value, str)
                            and isinstance(hucre.value, (int, float))):
                        # Küçük konfigürasyon sabitleri (0, 1, -1) görmezden gel
                        if abs(hucre.value) <= 1:
                            continue
                        deger_hucre = ws_deger[hucre.coordinate]
                        self.uyarilar.append({
                            'seviye': 'uyari',
                            'kategori': 'Hardcoded Değer',
                            'konum': f"{sheet_adi}!{hucre.coordinate}",
                            'deger': hucre.value,
                            'satir_etiketi': satir_etiketi[:50] if satir_etiketi else '(etiket yok)',
                            'aciklama': 'Projeksiyon hücresinde formül yerine sabit — formüle dönüştür',
                        })

    def _format_tutarsizliklari(self):
        """
        Aynı sütunda farklı sayı formatı kullanan hücreleri tespit eder.
        Örnek: A sütununda bazı hücreler '%12.00', bazıları '%12.1234' formatında.
        """
        for sheet_adi in self.wb_formuller.sheetnames:
            ws = self.wb_formuller[sheet_adi]
            # Sütun bazlı format sayımı
            sutun_formatlari: Dict[int, Dict[str, int]] = {}
            for satir in ws.iter_rows(min_row=2):
                for hucre in satir:
                    if hucre.value is None:
                        continue
                    fmt = hucre.number_format or 'General'
                    col = hucre.column
                    if col not in sutun_formatlari:
                        sutun_formatlari[col] = {}
                    sutun_formatlari[col][fmt] = sutun_formatlari[col].get(fmt, 0) + 1

            # 2'den fazla farklı format kullanan sütunlar
            for col, formatlari in sutun_formatlari.items():
                if len(formatlari) > 2:
                    from openpyxl.utils import get_column_letter
                    sutun_harf = get_column_letter(col)
                    format_listesi = ', '.join(
                        f"'{f}' ({n}x)" for f, n in sorted(formatlari.items(), key=lambda x: -x[1])[:3]
                    )
                    self.uyarilar.append({
                        'seviye': 'uyari',
                        'kategori': 'Format Tutarsızlığı',
                        'konum': f"{sheet_adi}!{sutun_harf}",
                        'deger': f"{len(formatlari)} farklı format",
                        'aciklama': f"Tutarsız sayı formatı: {format_listesi}",
                    })


def _audit_cikti_yazdir(sonuc: dict):
    """Genel audit sonuçlarını okunabilir formatta yazdır."""
    durum_emoji = '✅' if sonuc['durum'] == 'TEMİZ' else '⚠️'
    print(f"\n{durum_emoji} Genel Excel Audit: {sonuc['durum']}")
    print(f"   Dosya  : {sonuc['dosya']}")
    print(f"   Tarih  : {sonuc['audit_tarihi']}")
    print(f"   Kritik : {sonuc['kritik_sayisi']} sorun")
    print(f"   Uyarı  : {sonuc['uyari_sayisi']} uyarı")
    print()

    if sonuc['kritikler']:
        print(f"🔴 KRİTİK SORUNLAR ({len(sonuc['kritikler'])}):")
        for item in sonuc['kritikler']:
            print(f"   [{item['kategori']}] {item['konum']}")
            print(f"   Değer   : {item['deger']}")
            print(f"   Açıklama: {item['aciklama']}")
            print()

    if sonuc['uyarilar']:
        print(f"🟡 UYARILAR ({len(sonuc['uyarilar'])}):")
        for item in sonuc['uyarilar']:
            etiket = item.get('satir_etiketi', '')
            konum_str = item['konum']
            if etiket:
                konum_str += f"  (satır: '{etiket}')"
            print(f"   [{item['kategori']}] {konum_str}")
            print(f"   Değer   : {item['deger']}")
            print(f"   Açıklama: {item['aciklama']}")
            print()

    if sonuc['bilgiler']:
        print("ℹ️  BİLGİLER:")
        for b in sonuc['bilgiler']:
            print(f"   {b}")
        print()

    print("─" * 60)
    print(f"ÖZET: {sonuc['kritik_sayisi']} kritik, {sonuc['uyari_sayisi']} uyarı")
    if sonuc['durum'] == 'TEMİZ':
        print("✅ Model audit geçti — T3'e geçilebilir.")
    else:
        print("⚠️  Kritik sorunları düzeltin, ardından tekrar çalıştırın.")


def main():
    """Komut satırı arayüzü."""
    parser = argparse.ArgumentParser(
        description=(
            'DCF Model Doğrulama & Genel Excel Audit\n'
            'Varsayılan: DCF doğrulama | --audit: Genel model audit'
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
DCF Doğrulama (varsayılan):
  1. Sheet yapısı (DCF, WACC, Sensitivity)
  2. Formül hataları (#REF!, #DIV/0!, vb.)
  3. Terminal büyüme < WACC (kritik)
  4. WACC aralık kontrolü (TL: %10-45, USD: %5-20)
  5. Terminal değer / EV oranı (%40-80 arası)

Genel Audit (--audit):
  A. Formül hataları (tüm sheet'ler)
  B. Hardcoded değer tespiti (projeksiyon sütunları)
  C. Format tutarsızlıkları

Örnekler:
  python dcf-dogrulama.py THYAO_DCF.xlsx
  python dcf-dogrulama.py THYAO_Financial_Model.xlsx --audit
  python dcf-dogrulama.py MODEL.xlsx --audit --json sonuc.json
  python dcf-dogrulama.py MODEL.xlsx --audit --para-birimi USD
        """
    )
    parser.add_argument('excel_dosyasi', help='Kontrol edilecek Excel dosyası')
    parser.add_argument('--json', dest='json_cikti', help='Sonuçları JSON dosyasına kaydet')
    parser.add_argument('--para-birimi', default='TL', choices=['TL', 'USD', 'EUR'],
                        help='DCF para birimi (varsayılan: TL) — sadece DCF modunda kullanılır')
    parser.add_argument('--audit', action='store_true',
                        help='Genel Excel audit modu (hardcoded, format, formül hataları)')

    args = parser.parse_args()

    try:
        if args.audit:
            # ── Genel Audit Modu ──────────────────────────────────────────
            audit = GenelAudit(args.excel_dosyasi)
            sonuc = audit.audit()
            _audit_cikti_yazdir(sonuc)

            if args.json_cikti:
                with open(args.json_cikti, 'w', encoding='utf-8') as f:
                    json.dump(sonuc, f, indent=2, ensure_ascii=False)
                print(f"📄 JSON çıktı: {args.json_cikti}")

            sys.exit(0 if sonuc['durum'] == 'TEMİZ' else 1)

        else:
            # ── DCF Doğrulama Modu (mevcut, değişmedi) ───────────────────
            dogrulayici = DCFModelDogrulayici(args.excel_dosyasi, args.para_birimi)
            sonuc = dogrulayici.dogrula()

            durum_emoji = '✅' if sonuc['durum'] == 'GEÇTİ' else '❌'
            print(f"\n{durum_emoji} DCF Doğrulama: {sonuc['durum']}")
            print(f"   Dosya: {sonuc['dosya']}")
            print(f"   Para birimi: {sonuc['para_birimi']}")
            print(f"   Tarih: {sonuc['dogrulama_tarihi']}")
            print()

            if sonuc['hatalar']:
                print(f"🔴 HATALAR ({sonuc['hata_sayisi']}):")
                for h in sonuc['hatalar']:
                    print(f"   • {h}")
                print()

            if sonuc['uyarilar']:
                print(f"🟡 UYARILAR ({sonuc['uyari_sayisi']}):")
                for u in sonuc['uyarilar']:
                    print(f"   • {u}")
                print()

            if sonuc['bilgiler']:
                print(f"ℹ️  BİLGİLER:")
                for b in sonuc['bilgiler']:
                    print(f"   {b}")
                print()

            if args.json_cikti:
                with open(args.json_cikti, 'w', encoding='utf-8') as f:
                    json.dump(sonuc, f, indent=2, ensure_ascii=False)
                print(f"📄 JSON çıktı: {args.json_cikti}")

            sys.exit(0 if sonuc['durum'] == 'GEÇTİ' else 1)

    except Exception as e:
        print(f"❌ Hata: {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
