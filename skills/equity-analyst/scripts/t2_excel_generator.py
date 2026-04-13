#!/usr/bin/env python3
"""
T2 Excel Model Generator — BBB Araştırma Standart Finansal Model
Kullanım: python3 t2_excel_generator.py TICKER [--output PATH]

Bu script BBB Finans'tan veri çeker ve standart 6-tab Excel modeli üretir.
Her şirket için aynı yapı, aynı formüller, aynı referanslar.
"""

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter
except ImportError:
    print("❌ openpyxl gerekli: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# STİL TANIMLARI (EBEBK/NVO standardı)
# =============================================================================
STYLES = {
    'TITLE': Font(color="2C3E50", bold=True, size=14),
    'SUBTITLE': Font(color="888888", bold=False, size=9),
    'HDR_FONT': Font(color="FFFFFF", bold=True, size=11),
    'HDR_FILL': PatternFill("solid", fgColor="2C3E50"),
    'MAVI': Font(color="0000FF", size=10),  # Input
    'SIYAH': Font(color="000000", size=10),  # Normal
    'SIYAH_BOLD': Font(color="000000", bold=True, size=10),
    'YESIL': Font(color="008000", size=10),  # Cross-sheet link
    'KIRMIZI': Font(color="FF0000", bold=True, size=10),  # Warning
    'ITALIK': Font(color="666666", italic=True, size=9),
    'GRI_FILL': PatternFill("solid", fgColor="E8E8E8"),  # Tarihsel
    'BOLUM_FILL': PatternFill("solid", fgColor="D5E8D4"),  # Bölüm başlık
    'TOPLAM_FILL': PatternFill("solid", fgColor="D6E4F0"),  # Toplam satır
    'SENARYO_FILL': PatternFill("solid", fgColor="FFFFCC"),  # Input/Senaryo
}

THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)

# Tab renkleri
TAB_COLORS = {
    'GelirModeli': 'F7931A',  # Turuncu
    'GelirTablosu': 'E74C3C',  # Kırmızı
    'NakitAkis': '27AE60',  # Yeşil
    'Bilanco': '3498DB',  # Mavi
    'Senaryolar': '9B59B6',  # Mor
    'INAGirdileri': '34495E',  # Koyu gri
}

# Yıllar
HISTORICAL_YEARS = ['2021/FY', '2022/FY', '2023/FY', '2024/FY', '2025/FY']
PROJECTION_YEARS = ['FY2026T', 'FY2027T']
ALL_COLS = ['FY2021', 'FY2022', 'FY2023', 'FY2024', 'FY2025'] + PROJECTION_YEARS


# =============================================================================
# VERİ ÇEKME
# =============================================================================
def fetch_financial_data(ticker: str) -> dict:
    """BBB Finans'tan finansal veri çeker."""
    scripts_dir = Path(__file__).parent.parent.parent / 'bbb-finans' / 'scripts'
    
    cmd = [
        'python3', str(scripts_dir / 'bbb_financials.py'),
        ticker, '--dcf', '--json'
    ]
    
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if result.returncode != 0:
            print(f"❌ BBB Finans hatası: {result.stderr}")
            sys.exit(1)
        return json.loads(result.stdout)
    except subprocess.TimeoutExpired:
        print("❌ BBB Finans zaman aşımı")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"❌ JSON parse hatası: {e}")
        sys.exit(1)


def get_value(data: dict, code: str, period: str, divisor: float = 1e9) -> float:
    """Belirli bir kalem ve dönem için değer döndürür."""
    if code not in data.get('items', {}):
        return 0.0
    
    val = data['items'][code].get('all_periods', {}).get(period, 0)
    return (val or 0) / divisor


def get_all_years(data: dict, code: str) -> list:
    """Tüm tarihsel yıllar için değerleri döndürür."""
    return [get_value(data, code, y) for y in HISTORICAL_YEARS]


# =============================================================================
# SAYFA KURULUM
# =============================================================================
def setup_sheet(ws, title: str, subtitle1: str, subtitle2: str, 
                headers: list, tab_color: str = None):
    """Standart sayfa kurulumu."""
    ncol = len(headers)
    
    # Sütun genişlikleri
    ws.column_dimensions['A'].width = 42
    for i in range(2, ncol + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    
    # Freeze ve görünüm
    ws.freeze_panes = 'B6'
    ws.sheet_view.showGridLines = False
    
    if tab_color:
        ws.sheet_properties.tabColor = tab_color
    
    # Başlıklar
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
    ws.cell(row=1, column=1, value=title).font = STYLES['TITLE']
    ws.cell(row=2, column=1, value=subtitle1).font = STYLES['SUBTITLE']
    ws.cell(row=3, column=1, value=subtitle2).font = STYLES['SUBTITLE']
    
    # Header row
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=i, value=h)
        c.font = STYLES['HDR_FONT']
        c.fill = STYLES['HDR_FILL']
        c.alignment = Alignment(horizontal='center')
        c.border = THIN_BORDER


def write_section_header(ws, row: int, title: str) -> int:
    """Bölüm başlığı yazar."""
    ws.cell(row=row, column=1, value=title).font = STYLES['SIYAH_BOLD']
    ws.cell(row=row, column=1).fill = STYLES['BOLUM_FILL']
    return row + 1


def write_data_row(ws, row: int, label: str, values: list, 
                   is_total: bool = False, is_input: bool = False,
                   number_format: str = '#,##0.0') -> int:
    """Veri satırı yazar (tarihsel + projeksiyon)."""
    
    # Label
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = STYLES['SIYAH_BOLD'] if is_total else STYLES['SIYAH']
    
    if is_total:
        ws.cell(row=row, column=1).fill = STYLES['TOPLAM_FILL']
    
    # Tarihsel değerler (5 yıl)
    for i, val in enumerate(values[:5]):
        c = ws.cell(row=row, column=i + 2, value=val)
        c.number_format = number_format
        c.font = STYLES['SIYAH_BOLD'] if is_total else STYLES['SIYAH']
        c.fill = STYLES['TOPLAM_FILL'] if is_total else STYLES['GRI_FILL']
    
    # Projeksiyon değerleri (2 yıl) - placeholder
    for col in [7, 8]:
        c = ws.cell(row=row, column=col, value=values[-1] if values else 0)
        c.number_format = number_format
        c.font = STYLES['MAVI'] if is_input else STYLES['YESIL']
        if is_total:
            c.fill = STYLES['TOPLAM_FILL']
    
    return row + 1


def write_formula_row(ws, row: int, label: str, formula_template: str,
                      is_total: bool = False, number_format: str = '#,##0.0') -> int:
    """Formül satırı yazar."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = STYLES['SIYAH_BOLD'] if is_total else STYLES['ITALIK']
    
    if is_total:
        ws.cell(row=row, column=1).fill = STYLES['TOPLAM_FILL']
    
    for i, col_letter in enumerate(['B', 'C', 'D', 'E', 'F', 'G', 'H']):
        formula = formula_template.replace('{COL}', col_letter)
        c = ws.cell(row=row, column=i + 2, value=formula)
        c.number_format = number_format
        c.font = STYLES['SIYAH_BOLD'] if is_total else STYLES['ITALIK']
        
        if is_total:
            c.fill = STYLES['TOPLAM_FILL']
        elif i < 5:
            c.fill = STYLES['GRI_FILL']
    
    return row + 1


# =============================================================================
# SAYFA ÜRETİCİLER
# =============================================================================
class RowTracker:
    """Dinamik satır takibi için yardımcı sınıf."""
    def __init__(self):
        self.rows = {}
    
    def set(self, name: str, row: int):
        self.rows[name] = row
    
    def get(self, name: str) -> int:
        return self.rows.get(name, 0)


def create_gelir_modeli(wb, data: dict, ticker: str, tracker: RowTracker):
    """GelirModeli sayfası oluşturur."""
    ws = wb.active
    ws.title = "GelirModeli"
    
    headers = ['Kalem'] + ALL_COLS
    setup_sheet(ws, f"{ticker} — Gelir Modeli", 
                "Para birimi: Milyar TL | Kaynak: BBB Finans (İş Yatırım)",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['GelirModeli'])
    
    hasilat = get_all_years(data, '3C')
    row = 7
    
    # Bölüm A: Toplam Hasılat
    row = write_section_header(ws, row, "BÖLÜM A: TOPLAM HASILAT")
    
    # Büyüme varsayımı
    ws.cell(row=row, column=1, value="YoY Büyüme Varsayımı (%)").font = STYLES['MAVI']
    for i in range(1, 5):
        if hasilat[i-1] > 0:
            growth = hasilat[i] / hasilat[i-1] - 1
        else:
            growth = 0
        c = ws.cell(row=row, column=i + 2, value=growth)
        c.number_format = '0.0%'
        c.font = STYLES['SIYAH']
        c.fill = STYLES['GRI_FILL']
    
    # Projeksiyon büyüme - Senaryolar'a bağlı
    c = ws.cell(row=row, column=7, value='=IF(Senaryolar!$B$5=1,Senaryolar!B7,IF(Senaryolar!$B$5=3,Senaryolar!D7,Senaryolar!C7))')
    c.number_format = '0.0%'
    c.font = STYLES['YESIL']
    c = ws.cell(row=row, column=8, value='=IF(Senaryolar!$B$5=1,Senaryolar!B8,IF(Senaryolar!$B$5=3,Senaryolar!D8,Senaryolar!C8))')
    c.number_format = '0.0%'
    c.font = STYLES['YESIL']
    
    tracker.set('buyume_row', row)
    row += 1
    
    # Toplam Hasılat
    ws.cell(row=row, column=1, value="TOPLAM HASILAT").font = STYLES['SIYAH_BOLD']
    ws.cell(row=row, column=1).fill = STYLES['TOPLAM_FILL']
    
    for i, val in enumerate(hasilat):
        c = ws.cell(row=row, column=i + 2, value=val)
        c.number_format = '#,##0.0'
        c.font = STYLES['SIYAH_BOLD']
        c.fill = STYLES['TOPLAM_FILL']
    
    # Projeksiyon formülleri
    buyume_row = tracker.get('buyume_row')
    c = ws.cell(row=row, column=7, value=f'=F{row}*(1+G{buyume_row})')
    c.number_format = '#,##0.0'
    c.font = STYLES['YESIL']
    c.fill = STYLES['TOPLAM_FILL']
    
    c = ws.cell(row=row, column=8, value=f'=G{row}*(1+H{buyume_row})')
    c.number_format = '#,##0.0'
    c.font = STYLES['YESIL']
    c.fill = STYLES['TOPLAM_FILL']
    
    tracker.set('hasilat_row', row)
    row += 2
    
    # Bölüm B: Segment Payları
    row = write_section_header(ws, row, "BÖLÜM B: SEGMENT PAYLARI (input)")
    
    # Örnek segment payları - her şirket için özelleştirilmeli
    segments = [
        ("Segment 1 Pay (%)", 0.60),
        ("Segment 2 Pay (%)", 0.25),
        ("Segment 3 Pay (%)", 0.10),
        ("Diğer Pay (%)", 0.05),
    ]
    
    hasilat_row = tracker.get('hasilat_row')
    for seg_name, seg_val in segments:
        ws.cell(row=row, column=1, value=seg_name).font = STYLES['MAVI']
        for col in range(2, 9):
            c = ws.cell(row=row, column=col, value=seg_val)
            c.number_format = '0.0%'
            c.font = STYLES['MAVI']
            c.fill = STYLES['GRI_FILL'] if col < 7 else STYLES['SENARYO_FILL']
        row += 1
        
        # Segment geliri
        seg_gelir_name = seg_name.replace("Pay (%)", "Gelir")
        ws.cell(row=row, column=1, value=f"  {seg_gelir_name}").font = STYLES['SIYAH']
        for col in range(2, 9):
            col_letter = get_column_letter(col)
            c = ws.cell(row=row, column=col, value=f'={col_letter}{hasilat_row}*{col_letter}{row-1}')
            c.number_format = '#,##0.0'
            c.font = STYLES['SIYAH']
            c.fill = STYLES['GRI_FILL'] if col < 7 else PatternFill()
        row += 1
    
    row += 1
    
    # Bölüm C: Coğrafi Kırılım
    row = write_section_header(ws, row, "BÖLÜM C: COĞRAFİ KIRILIM")
    
    ws.cell(row=row, column=1, value="Türkiye Payı (%)").font = STYLES['MAVI']
    for col in range(2, 9):
        c = ws.cell(row=row, column=col, value=0.99)
        c.number_format = '0.0%'
        c.font = STYLES['MAVI']
        c.fill = STYLES['GRI_FILL'] if col < 7 else STYLES['SENARYO_FILL']
    tracker.set('tr_pay_row', row)
    row += 1
    
    ws.cell(row=row, column=1, value="  Türkiye Gelir").font = STYLES['SIYAH']
    tr_pay_row = tracker.get('tr_pay_row')
    for col in range(2, 9):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f'={col_letter}{hasilat_row}*{col_letter}{tr_pay_row}')
        c.number_format = '#,##0.0'
        c.font = STYLES['SIYAH']
    row += 2
    
    # CRP
    ws.cell(row=row, column=1, value="Türkiye CRP (%)").font = STYLES['SIYAH']
    for col in range(2, 9):
        c = ws.cell(row=row, column=col, value=0.0425)
        c.number_format = '0.0%'
        c.font = STYLES['MAVI']
    
    return ws


def create_gelir_tablosu(wb, data: dict, ticker: str, tracker: RowTracker):
    """GelirTablosu sayfası oluşturur."""
    ws = wb.create_sheet("GelirTablosu")
    
    headers = ['Kalem'] + ALL_COLS
    setup_sheet(ws, f"{ticker} — Gelir Tablosu", 
                "Para birimi: Milyar TL | Kaynak: BBB Finans (İş Yatırım)",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['GelirTablosu'])
    
    row = 7
    hasilat_row_gm = tracker.get('hasilat_row')
    
    # Gelir Tablosu bölümü
    row = write_section_header(ws, row, "GELİR TABLOSU")
    
    # Hasılat - GelirModeli'nden link
    ws.cell(row=row, column=1, value="Satış Gelirleri (Hasılat)").font = STYLES['SIYAH_BOLD']
    ws.cell(row=row, column=1).fill = STYLES['TOPLAM_FILL']
    
    hasilat = get_all_years(data, '3C')
    for i, val in enumerate(hasilat):
        c = ws.cell(row=row, column=i + 2, value=val)
        c.number_format = '#,##0.0'
        c.font = STYLES['SIYAH_BOLD']
        c.fill = STYLES['TOPLAM_FILL']
    
    for col in [7, 8]:
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f'=GelirModeli!{col_letter}{hasilat_row_gm}')
        c.number_format = '#,##0.0'
        c.font = STYLES['YESIL']
        c.fill = STYLES['TOPLAM_FILL']
    
    tracker.set('hasilat_row_gt', row)
    row += 1
    
    # SMM
    row = write_data_row(ws, row, "Satışların Maliyeti (-)", get_all_years(data, '3CA'))
    tracker.set('smm_row', row - 1)
    
    # Brüt Kâr
    hasilat_row_gt = tracker.get('hasilat_row_gt')
    smm_row = tracker.get('smm_row')
    row = write_formula_row(ws, row, "BRÜT KÂR", f'={{COL}}{hasilat_row_gt}+{{COL}}{smm_row}', is_total=True)
    tracker.set('brut_kar_row', row - 1)
    
    # Brüt Kâr Marjı
    brut_kar_row = tracker.get('brut_kar_row')
    row = write_formula_row(ws, row, "  Brüt Kâr Marjı (%)", f'={{COL}}{brut_kar_row}/{{COL}}{hasilat_row_gt}', number_format='0.0%')
    row += 1
    
    # Faaliyet Giderleri
    row = write_section_header(ws, row, "FAALİYET GİDERLERİ")
    
    row = write_data_row(ws, row, "Pazarlama Giderleri (-)", get_all_years(data, '3DA'))
    tracker.set('paz_row', row - 1)
    
    row = write_data_row(ws, row, "Genel Yönetim Giderleri (-)", get_all_years(data, '3DB'))
    tracker.set('gyg_row', row - 1)
    
    row = write_data_row(ws, row, "Ar-Ge Giderleri (-)", get_all_years(data, '3DC'))
    tracker.set('arge_row', row - 1)
    
    row = write_data_row(ws, row, "Diğer Faaliyet Gelir/Gider", [5, 5, 5, 5, 5])
    tracker.set('diger_faal_row', row - 1)
    row += 1
    
    # EBIT
    paz_row = tracker.get('paz_row')
    gyg_row = tracker.get('gyg_row')
    arge_row = tracker.get('arge_row')
    diger_faal_row = tracker.get('diger_faal_row')
    
    row = write_formula_row(ws, row, "FAALİYET KÂRI (EBIT)", 
        f'={{COL}}{brut_kar_row}+{{COL}}{paz_row}+{{COL}}{gyg_row}+{{COL}}{arge_row}+{{COL}}{diger_faal_row}',
        is_total=True)
    tracker.set('ebit_row', row - 1)
    
    # EBIT Marjı
    ebit_row = tracker.get('ebit_row')
    row = write_formula_row(ws, row, "  EBIT Marjı (%)", f'={{COL}}{ebit_row}/{{COL}}{hasilat_row_gt}', number_format='0.0%')
    row += 1
    
    # D&A
    row = write_data_row(ws, row, "Amortisman & İtfa", get_all_years(data, '4B'))
    tracker.set('da_row', row - 1)
    
    # FAVÖK
    da_row = tracker.get('da_row')
    row = write_formula_row(ws, row, "FAVÖK (EBITDA)", f'={{COL}}{ebit_row}+{{COL}}{da_row}', is_total=True)
    tracker.set('favok_row', row - 1)
    
    # FAVÖK Marjı
    favok_row = tracker.get('favok_row')
    row = write_formula_row(ws, row, "  FAVÖK Marjı (%)", f'={{COL}}{favok_row}/{{COL}}{hasilat_row_gt}', number_format='0.0%')
    row += 1
    
    # Finansman
    row = write_section_header(ws, row, "FİNANSMAN")
    
    row = write_data_row(ws, row, "Finansal Gelirler", get_all_years(data, '3HB'))
    tracker.set('fin_gelir_row', row - 1)
    
    row = write_data_row(ws, row, "Finansal Giderler (-)", get_all_years(data, '3HC'))
    tracker.set('fin_gider_row', row - 1)
    row += 1
    
    # VÖK
    fin_gelir_row = tracker.get('fin_gelir_row')
    fin_gider_row = tracker.get('fin_gider_row')
    row = write_formula_row(ws, row, "VERGİ ÖNCESİ KÂR", 
        f'={{COL}}{ebit_row}+{{COL}}{fin_gelir_row}+{{COL}}{fin_gider_row}',
        is_total=True)
    tracker.set('vok_row', row - 1)
    
    # Vergi
    row = write_data_row(ws, row, "Vergi Gideri (-)", get_all_years(data, '3IA'))
    tracker.set('vergi_row', row - 1)
    
    # Net Kâr
    vok_row = tracker.get('vok_row')
    vergi_row = tracker.get('vergi_row')
    row = write_formula_row(ws, row, "NET KÂR", f'={{COL}}{vok_row}+{{COL}}{vergi_row}', is_total=True)
    tracker.set('net_kar_row', row - 1)
    
    # Net Kâr Marjı
    net_kar_row = tracker.get('net_kar_row')
    row = write_formula_row(ws, row, "  Net Kâr Marjı (%)", f'={{COL}}{net_kar_row}/{{COL}}{hasilat_row_gt}', number_format='0.0%')
    
    return ws


def create_nakit_akis(wb, data: dict, ticker: str, tracker: RowTracker):
    """NakitAkis sayfası oluşturur."""
    ws = wb.create_sheet("NakitAkis")
    
    headers = ['Kalem'] + ALL_COLS
    setup_sheet(ws, f"{ticker} — Nakit Akış Tablosu", 
                "Para birimi: Milyar TL | Kaynak: BBB Finans (İş Yatırım)",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['NakitAkis'])
    
    row = 7
    favok_row = tracker.get('favok_row')
    hasilat_row_gm = tracker.get('hasilat_row')
    
    # İşletme Faaliyetleri
    row = write_section_header(ws, row, "İŞLETME FAALİYETLERİ")
    
    # FAVÖK link
    ws.cell(row=row, column=1, value="FAVÖK").font = STYLES['SIYAH']
    for col in range(2, 9):
        col_letter = get_column_letter(col)
        c = ws.cell(row=row, column=col, value=f'=GelirTablosu!{col_letter}{favok_row}')
        c.number_format = '#,##0.0'
        c.font = STYLES['YESIL']
    tracker.set('favok_cf_row', row)
    row += 1
    
    row = write_data_row(ws, row, "(-) Vergi Ödemesi", get_all_years(data, '3IA'))
    tracker.set('vergi_cf_row', row - 1)
    
    row = write_data_row(ws, row, "(-) Net Faiz Ödemesi", get_all_years(data, '4BB'))
    tracker.set('faiz_cf_row', row - 1)
    
    row = write_data_row(ws, row, "İşletme Sermayesi Değişimi", get_all_years(data, '4CAF'), is_input=True)
    tracker.set('is_row', row - 1)
    
    # CFO
    favok_cf_row = tracker.get('favok_cf_row')
    vergi_cf_row = tracker.get('vergi_cf_row')
    faiz_cf_row = tracker.get('faiz_cf_row')
    is_row = tracker.get('is_row')
    
    row = write_formula_row(ws, row, "İşletme Faaliyetlerinden Nakit (CFO)",
        f'={{COL}}{favok_cf_row}+{{COL}}{vergi_cf_row}+{{COL}}{faiz_cf_row}+{{COL}}{is_row}',
        is_total=True)
    tracker.set('cfo_row', row - 1)
    
    # CFO Marjı
    cfo_row = tracker.get('cfo_row')
    row = write_formula_row(ws, row, "  CFO Marjı (%)", 
        f'={{COL}}{cfo_row}/GelirModeli!{{COL}}{hasilat_row_gm}', number_format='0.0%')
    row += 1
    
    # Yatırım Faaliyetleri
    row = write_section_header(ws, row, "YATIRIM FAALİYETLERİ")
    
    row = write_data_row(ws, row, "CapEx", get_all_years(data, '4CAI'))
    tracker.set('capex_row', row - 1)
    
    # CapEx/Hasılat
    capex_row = tracker.get('capex_row')
    row = write_formula_row(ws, row, "  CapEx/Hasılat (%)", 
        f'=-{{COL}}{capex_row}/GelirModeli!{{COL}}{hasilat_row_gm}', number_format='0.0%')
    row += 1
    
    # FCF
    row = write_formula_row(ws, row, "SERBEST NAKİT AKIŞI (FCF)",
        f'={{COL}}{cfo_row}+{{COL}}{capex_row}', is_total=True)
    tracker.set('fcf_row', row - 1)
    
    # FCF Marjı
    fcf_row = tracker.get('fcf_row')
    row = write_formula_row(ws, row, "  FCF Marjı (%)", 
        f'={{COL}}{fcf_row}/GelirModeli!{{COL}}{hasilat_row_gm}', number_format='0.0%')
    
    return ws


def create_bilanco(wb, data: dict, ticker: str, tracker: RowTracker):
    """Bilanco sayfası oluşturur."""
    ws = wb.create_sheet("Bilanco")
    
    headers = ['Kalem'] + ALL_COLS
    setup_sheet(ws, f"{ticker} — Bilanço", 
                "Para birimi: Milyar TL | Kaynak: BBB Finans (İş Yatırım)",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['Bilanco'])
    
    row = 7
    
    # Dönen Varlıklar
    row = write_section_header(ws, row, "DÖNEN VARLIKLAR")
    
    row = write_data_row(ws, row, "  Nakit ve Nakit Benzerleri", get_all_years(data, '1AA'))
    tracker.set('nakit_row', row - 1)
    
    row = write_data_row(ws, row, "  Ticari Alacaklar", get_all_years(data, '1AC'))
    row = write_data_row(ws, row, "  Stoklar", get_all_years(data, '1AF'))
    
    row = write_data_row(ws, row, "Toplam Dönen Varlıklar", get_all_years(data, '1A'), is_total=True)
    tracker.set('dv_row', row - 1)
    row += 1
    
    # Duran Varlıklar
    row = write_section_header(ws, row, "DURAN VARLIKLAR")
    row = write_data_row(ws, row, "Toplam Duran Varlıklar", [0]*5, is_total=True)
    tracker.set('dur_row', row - 1)
    row += 1
    
    # Toplam Varlıklar
    row = write_data_row(ws, row, "TOPLAM VARLIKLAR", get_all_years(data, '1BL'), is_total=True)
    tracker.set('tv_row', row - 1)
    row += 1
    
    # KV Yükümlülükler
    row = write_section_header(ws, row, "KISA VADELİ YÜKÜMLÜLÜKLER")
    
    row = write_data_row(ws, row, "  KV Finansal Borçlar", get_all_years(data, '2AA'))
    tracker.set('kv_borc_row', row - 1)
    
    row = write_data_row(ws, row, "  Ticari Borçlar", get_all_years(data, '2AAGAA'))
    
    row = write_data_row(ws, row, "Toplam KV Yükümlülükler", get_all_years(data, '2A'), is_total=True)
    tracker.set('kv_row', row - 1)
    row += 1
    
    # UV Yükümlülükler
    row = write_section_header(ws, row, "UZUN VADELİ YÜKÜMLÜLÜKLER")
    
    row = write_data_row(ws, row, "  UV Finansal Borçlar", get_all_years(data, '2BA'))
    tracker.set('uv_borc_row', row - 1)
    
    row = write_data_row(ws, row, "Toplam UV Yükümlülükler", [0]*5, is_total=True)
    tracker.set('uv_row', row - 1)
    row += 1
    
    # Özkaynaklar
    row = write_section_header(ws, row, "ÖZKAYNAKLAR")
    
    row = write_data_row(ws, row, "  Ödenmiş Sermaye", get_all_years(data, '2OA'))
    row = write_data_row(ws, row, "Ana Ortaklığa Ait Özkaynaklar", get_all_years(data, '2O'), is_total=True)
    row = write_data_row(ws, row, "TOPLAM ÖZKAYNAKLAR", get_all_years(data, '2N'), is_total=True)
    tracker.set('ok_row', row - 1)
    row += 1
    
    # Finansal Oranlar
    row = write_section_header(ws, row, "FİNANSAL ORANLAR")
    
    nakit_row = tracker.get('nakit_row')
    kv_borc_row = tracker.get('kv_borc_row')
    uv_borc_row = tracker.get('uv_borc_row')
    favok_row = tracker.get('favok_row')
    
    row = write_formula_row(ws, row, "Net Finansal Borç",
        f'={{COL}}{kv_borc_row}+{{COL}}{uv_borc_row}-{{COL}}{nakit_row}')
    tracker.set('net_borc_row', row - 1)
    
    net_borc_row = tracker.get('net_borc_row')
    row = write_formula_row(ws, row, "Net Borç / FAVÖK",
        f'={{COL}}{net_borc_row}/GelirTablosu!{{COL}}{favok_row}', number_format='0.0x')
    
    return ws


def create_senaryolar(wb, ticker: str, tracker: RowTracker):
    """Senaryolar sayfası oluşturur."""
    ws = wb.create_sheet("Senaryolar")
    
    headers = ['FM Parametresi', 'Bear', 'Base', 'Bull', 'Not/Kaynak']
    setup_sheet(ws, f"{ticker} — Senaryo Varsayımları", 
                "T1 research + yönetim guidance + kurum konsensüsü",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['Senaryolar'])
    ws.column_dimensions['E'].width = 35
    
    # Senaryo seçici
    ws.cell(row=4, column=1, value="AKTİF SENARYO:").font = STYLES['KIRMIZI']
    ws.cell(row=4, column=2, value="Base").font = STYLES['KIRMIZI']
    ws.cell(row=4, column=2).fill = STYLES['SENARYO_FILL']
    
    # DataValidation
    dv = DataValidation(type="list", formula1='"Bear,Base,Bull"', allow_blank=False)
    ws.add_data_validation(dv)
    dv.add(ws['B4'])
    
    # Senaryo kodu
    ws.cell(row=5, column=1, value="Senaryo Kodu (otomatik)").font = STYLES['ITALIK']
    ws.cell(row=5, column=2, value='=IF(B4="Bear",1,IF(B4="Bull",3,2))').font = STYLES['ITALIK']
    
    # Parametreler
    row = 7
    params = [
        ("YoY Büyüme FY2026T", 0.03, 0.06, 0.10, "Guidance"),
        ("YoY Büyüme FY2027T", 0.03, 0.05, 0.08, "Normalize"),
        ("Terminal Büyüme (g)", 0.025, 0.035, 0.045, "Reel GDP"),
        ("Terminal FAVÖK Marjı", 0.35, 0.40, 0.45, "Guidance"),
        ("ETR (Terminal)", 0.28, 0.25, 0.22, "Yasal oran"),
        ("CapEx/Hasılat FY26", 0.20, 0.18, 0.16, "Tarihsel"),
        ("İS Değişimi FY26 (B TL)", -15, -10, -5, "Büyümeyle artar"),
        ("İS Değişimi FY27 (B TL)", -15, -10, -5, ""),
        ("Senaryo Olasılığı", 0.25, 0.50, 0.25, "Varsayılan"),
    ]
    
    for param, bear, base, bull, note in params:
        ws.cell(row=row, column=1, value=param).font = STYLES['SIYAH']
        
        for col, val in [(2, bear), (3, base), (4, bull)]:
            c = ws.cell(row=row, column=col, value=val)
            c.font = STYLES['MAVI']
            c.fill = STYLES['SENARYO_FILL']
            
            if any(x in param for x in ['%', 'Büyüme', 'Marj', 'ETR', '/']):
                c.number_format = '0.0%'
            else:
                c.number_format = '#,##0.0'
        
        ws.cell(row=row, column=5, value=note).font = STYLES['ITALIK']
        row += 1
    
    return ws


def create_ina_girdileri(wb, ticker: str, tracker: RowTracker):
    """INAGirdileri sayfası oluşturur."""
    ws = wb.create_sheet("INAGirdileri")
    
    headers = ['Parametre', 'Değer', 'Kaynak/Not']
    setup_sheet(ws, f"{ticker} — İNA (DCF) Girdileri", 
                "bbb-dcf Faz 2 handoff",
                f"BBB Araştırma | {datetime.now().strftime('%Y-%m-%d')}",
                headers, TAB_COLORS['INAGirdileri'])
    ws.column_dimensions['C'].width = 40
    
    row = 7
    items = [
        ("BÜYÜME GİRDİLERİ", None, None, "bolum"),
        ("FY2026T Büyüme", "=Senaryolar!C7", "Senaryo", "yuzde"),
        ("FY2027T Büyüme", "=Senaryolar!C8", "Senaryo", "yuzde"),
        ("Terminal Büyüme (g)", "=Senaryolar!C9", "Senaryo", "yuzde"),
        ("", None, None, None),
        ("WACC GİRDİLERİ", None, None, "bolum"),
        ("Risksiz Getiri (Rf)", 0.045, "10Y US Treasury", "yuzde"),
        ("Olgunlaşmış Pazar ERP", 0.055, "Damodaran", "yuzde"),
        ("Ülke Risk Primi (CRP)", 0.0425, "Türkiye CDS", "yuzde"),
        ("Kaldıraçsız Beta (β_U)", 0.80, "Sektör ortalaması", "oran"),
        ("Vergi Oranı (t)", 0.25, "Yasal oran", "yuzde"),
        ("D/E Oranı", 0.30, "Bilanço'dan", "oran"),
        ("Borçlanma Maliyeti (Kd)", 0.08, "Sentetik/Gerçek", "yuzde"),
    ]
    
    for param, val, note, fmt in items:
        if not param:
            row += 1
            continue
        
        ws.cell(row=row, column=1, value=param)
        
        if fmt == "bolum":
            ws.cell(row=row, column=1).font = STYLES['SIYAH_BOLD']
            ws.cell(row=row, column=1).fill = STYLES['BOLUM_FILL']
        else:
            ws.cell(row=row, column=1).font = STYLES['SIYAH']
            
            if val is not None:
                c = ws.cell(row=row, column=2, value=val)
                c.alignment = Alignment(horizontal='right')
                c.font = STYLES['YESIL'] if isinstance(val, str) and val.startswith('=') else STYLES['MAVI']
                
                if fmt == "yuzde":
                    c.number_format = '0.0%'
                elif fmt == "oran":
                    c.number_format = '0.00'
        
        if note:
            ws.cell(row=row, column=3, value=note).font = STYLES['ITALIK']
        
        row += 1
    
    return ws


# =============================================================================
# ANA FONKSİYON
# =============================================================================
def generate_financial_model(ticker: str, output_path: str = None) -> str:
    """Standart finansal model Excel dosyası üretir."""
    
    print(f"📊 {ticker} finansal model üretiliyor...")
    
    # Veri çek
    print("  → BBB Finans'tan veri çekiliyor...")
    data = fetch_financial_data(ticker)
    
    if not data or 'items' not in data:
        print(f"❌ {ticker} için veri bulunamadı")
        sys.exit(1)
    
    print(f"  → {len(data['items'])} kalem bulundu")
    
    # Workbook oluştur
    wb = Workbook()
    tracker = RowTracker()
    
    # Sayfaları oluştur
    print("  → GelirModeli sayfası oluşturuluyor...")
    create_gelir_modeli(wb, data, ticker, tracker)
    
    print("  → GelirTablosu sayfası oluşturuluyor...")
    create_gelir_tablosu(wb, data, ticker, tracker)
    
    print("  → NakitAkis sayfası oluşturuluyor...")
    create_nakit_akis(wb, data, ticker, tracker)
    
    print("  → Bilanco sayfası oluşturuluyor...")
    create_bilanco(wb, data, ticker, tracker)
    
    print("  → Senaryolar sayfası oluşturuluyor...")
    create_senaryolar(wb, ticker, tracker)
    
    print("  → INAGirdileri sayfası oluşturuluyor...")
    create_ina_girdileri(wb, ticker, tracker)
    
    # Kaydet
    if output_path is None:
        output_path = f"{ticker}_Finansal_Model_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    wb.save(output_path)
    
    # Doğrulama
    from openpyxl import load_workbook as lw
    wb_check = lw(output_path)
    total_formulas = 0
    for sheet in wb_check.sheetnames:
        for row in wb_check[sheet].iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    total_formulas += 1
    
    print(f"\n✅ {ticker} finansal model oluşturuldu: {output_path}")
    print(f"   Sayfa sayısı: {len(wb_check.sheetnames)}")
    print(f"   Toplam formül: {total_formulas}")
    print(f"   Satır takibi: {len(tracker.rows)} referans")
    
    return output_path


def main():
    parser = argparse.ArgumentParser(description='T2 Excel Model Generator')
    parser.add_argument('ticker', help='Şirket ticker (örn: TCELL, THYAO)')
    parser.add_argument('--output', '-o', help='Çıktı dosya yolu')
    
    args = parser.parse_args()
    
    generate_financial_model(args.ticker, args.output)


if __name__ == '__main__':
    main()
