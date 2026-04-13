#!/usr/bin/env python3
"""
DCF Output Generator — Damodaran "Valuation as Picture" Exact Reproduction
Generates Excel and PDF outputs matching Damodaran's fcffsimpleginzu.xlsx layout.
"""

import os
import sys
from datetime import datetime
from dataclasses import dataclass, field
from typing import List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from copy import copy


# ──────────────────────────────────────────────────────────────────
# Data Classes
# ──────────────────────────────────────────────────────────────────

@dataclass
class DCFInput:
    """All inputs needed to generate Damodaran-style valuation output."""
    company_name: str = "Company"
    valuation_date: datetime = field(default_factory=datetime.now)
    currency_symbol: str = "$"

    # Base year
    base_revenue_growth: float = 0.05
    industry_revenue_growth: float = 0.07
    base_revenue: float = 10000
    base_operating_margin: float = 0.15
    industry_operating_margin: float = 0.12
    base_operating_income: float = 1500
    base_ebit_1_t: float = 1200

    # Stories
    growth_story: str = "Revenue growth driven by market expansion and new products."
    profitability_story: str = "Margins expected to improve with operating leverage."
    growth_efficiency_story: str = "Maintained at current reinvestment levels."
    risk_story: str = "Cost of capital reflects company risk profile."
    competitive_advantages: str = "Strong brand and market position."

    # 10-year projections (lists of 10)
    revenue_growth: List[float] = field(default_factory=lambda: [0.08]*5 + [0.06]*5)
    revenues: List[float] = field(default_factory=lambda: [10800, 11664, 12597, 13605, 14693, 15575, 16509, 17500, 18550, 19663])
    operating_margin: List[float] = field(default_factory=lambda: [0.15]*10)
    operating_income: List[float] = field(default_factory=lambda: [1620, 1750, 1890, 2041, 2204, 2336, 2476, 2625, 2783, 2949])
    ebit_1_t: List[float] = field(default_factory=lambda: [1296, 1400, 1512, 1633, 1763, 1869, 1981, 2100, 2226, 2360])
    reinvestment: List[float] = field(default_factory=lambda: [400, 432, 467, 504, 544, 441, 467, 496, 525, 557])
    fcff: List[float] = field(default_factory=lambda: [896, 968, 1045, 1129, 1219, 1428, 1514, 1604, 1701, 1803])
    cost_of_capital: List[float] = field(default_factory=lambda: [0.08]*5 + [0.085, 0.09, 0.095, 0.10, 0.10])
    cumulated_wacc: List[float] = field(default_factory=lambda: [0.926, 0.857, 0.794, 0.735, 0.680, 0.627, 0.575, 0.525, 0.477, 0.434])
    sales_to_capital: List[float] = field(default_factory=lambda: [2.0]*10)
    roic: List[float] = field(default_factory=lambda: [0.08, 0.085, 0.09, 0.095, 0.10, 0.10, 0.10, 0.10, 0.10, 0.10])

    # Terminal year
    terminal_revenue_growth: float = 0.04
    terminal_revenue: float = 20449
    terminal_operating_margin: float = 0.15
    terminal_operating_income: float = 3067
    terminal_ebit_1_t: float = 2454
    terminal_reinvestment: float = 982
    terminal_fcff: float = 1472
    terminal_cost_of_capital: float = 0.10
    terminal_roic: float = 0.10

    # Terminal value box
    tv_growth_rate: float = 0.04
    tv_cost_of_capital: float = 0.10
    tv_return_on_capital: float = 0.10
    tv_reinvestment_rate: float = 0.40

    # Equity bridge
    pv_terminal_value: float = 15000
    pv_cf_next_10: float = 8500
    probability_of_failure: float = 0.0
    value_of_operating_assets: float = 23500
    debt: float = 5000
    minority_interests: float = 0
    cash: float = 2000
    non_operating_assets: float = 500
    value_of_equity: float = 21000
    value_of_options: float = 0
    value_equity_common: float = 21000
    number_of_shares: float = 1000
    estimated_value_per_share: float = 21.00
    price_per_share: float = 18.50
    pct_under_over_valued: float = 0.135

    # Terminal value total (for display in projection area)
    terminal_value_total: float = 24533

    # Diagnostics
    industry_avg_cost_of_capital: float = 0.085
    compounded_cost_of_capital: float = 0.09
    failure_rate: float = 0.0
    industry_avg_revenue_growth: float = 0.07
    base_sales_to_capital: float = 1.5
    pv_ati_10yr: float = 12000
    pv_ati_pct: float = 1.0
    reinvestment_value_effect: float = 3500
    reinvestment_value_pct: float = 0.29
    pv_fcff_10yr: float = 8500
    pv_fcff_pct: float = 0.71
    roc_most_recent: float = 0.08
    roc_marginal: float = 0.12
    roc_year10: float = 0.10
    roc_stable: float = 0.10
    value_as_pct_of_price: float = 1.135


# ──────────────────────────────────────────────────────────────────
# Style Constants (matching Damodaran's format)
# ──────────────────────────────────────────────────────────────────

# Theme 2 in default Office theme is typically EEECE1 (light tan/cream)
# We'll use a light blue-gray that's common in Damodaran's sheets
LIGHT_FILL = PatternFill(patternType='solid', fgColor='DAEEF3')  # Light blue
STORY_FILL = PatternFill(patternType='solid', fgColor='DAEEF3')  # Same for story boxes
HIGHLIGHT_FILL = PatternFill(patternType='solid', fgColor='DAEEF3')

FONT_10 = Font(name='Calibri', size=10)
FONT_10_BOLD = Font(name='Calibri', size=10, bold=True)
FONT_12_BOLD = Font(name='Calibri', size=12, bold=True)
FONT_12 = Font(name='Calibri', size=12)

THIN = Side(style='thin')
MEDIUM = Side(style='medium')
NO_BORDER = Side(style=None)

BORDER_ALL_THIN = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_MEDIUM_BOX = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)

ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center')
ALIGN_RIGHT = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Number formats
NF_PCT = '0.00%'
NF_DOLLAR = '"$"#,##0'
NF_DOLLAR_DEC = '"$"#,##0.00'
NF_DOLLAR_ACCT = '_("$"* #,##0_);_("$"* \\(#,##0\\);_("$"* "-"??_);_(@_)'
NF_DOLLAR_ACCT_DEC = '_("$"* #,##0.00_);_("$"* \\(#,##0.00\\);_("$"* "-"??_);_(@_)'
NF_NUM_DEC = '0.00'
NF_NUM_4DEC = '0.0000'
NF_DATE = 'mmm-yy'
NF_SHARES = '_(* #,##0.00_);_(* \\(#,##0.00\\);_(* "-"??_);_(@_)'


def _set_cell(ws, row, col, value, font=None, fill=None, border=None, alignment=None, number_format=None):
    """Helper to set cell value and formatting."""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if border: cell.border = border
    if alignment: cell.alignment = alignment
    if number_format: cell.number_format = number_format
    return cell


def _apply_border_range(ws, min_row, max_row, min_col, max_col, border):
    """Apply border to a range of cells."""
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = border


def _col(letter):
    """Convert column letter to number."""
    from openpyxl.utils import column_index_from_string
    return column_index_from_string(letter)


def _make_currency_format(symbol):
    """Create number format with given currency symbol."""
    return f'"{symbol}"#,##0'


def _make_currency_format_dec(symbol):
    return f'"{symbol}"#,##0.00'


def _make_acct_format(symbol):
    return f'_("{symbol}"* #,##0_);_("{symbol}"* \\(#,##0\\);_("{symbol}"* "-"??_);_(@_)'


def _make_acct_format_dec(symbol):
    return f'_("{symbol}"* #,##0.00_);_("{symbol}"* \\(#,##0.00\\);_("{symbol}"* "-"??_);_(@_)'


# ──────────────────────────────────────────────────────────────────
# Valuation as Picture Sheet
# ──────────────────────────────────────────────────────────────────

def create_valuation_picture(wb, data: DCFInput):
    """Create the 'Valuation as picture' sheet — exact Damodaran layout."""
    ws = wb.create_sheet('Valuation as picture')
    
    sym = data.currency_symbol
    nf_dollar = _make_currency_format(sym)
    nf_dollar_dec = _make_currency_format_dec(sym)
    nf_acct = _make_acct_format(sym)
    nf_acct_dec = _make_acct_format_dec(sym)

    # Column widths (matching Damodaran)
    col_widths = {'A': 28, 'B': 13, 'C': 11, 'D': 16, 'E': 13, 'F': 11.5,
                  'G': 11, 'H': 11, 'I': 11, 'J': 11, 'K': 13, 'L': 11,
                  'M': 11, 'N': 14, 'O': 16, 'P': 11}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Row heights
    ws.row_dimensions[3].height = 17
    ws.row_dimensions[6].height = 14
    ws.row_dimensions[33].height = 14

    # ── Row 1: Instruction text ──
    _set_cell(ws, 1, 1, 
        "You can modify this picture and bring in relevant details to back up your cost of capital and other details that you think flesh out your company's valuation story.",
        font=FONT_10)

    # ── Row 3: Company name + Date ──
    ws.merge_cells('A3:N3')
    _set_cell(ws, 3, 1, data.company_name, font=FONT_12_BOLD)
    ws.merge_cells('O3:P3')
    _set_cell(ws, 3, _col('O'), data.valuation_date, font=FONT_10_BOLD, number_format=NF_DATE)

    # ── Row 5: Section headers ──
    # Base Year and Comparison
    ws.merge_cells('A5:C5')
    _set_cell(ws, 5, 1, 'Base Year and Comparison', font=FONT_10_BOLD,
              border=Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM))
    for c in [2, 3]:
        ws.cell(row=5, column=c).border = Border(top=MEDIUM, bottom=MEDIUM, left=THIN, right=MEDIUM if c==3 else THIN)
        ws.cell(row=5, column=c).font = FONT_10_BOLD

    # Growth Story
    ws.merge_cells('E5:F5')
    _set_cell(ws, 5, _col('E'), 'Growth Story', font=FONT_10_BOLD,
              border=Border(top=MEDIUM, bottom=NO_BORDER, left=MEDIUM, right=MEDIUM))
    ws.cell(row=5, column=_col('F')).border = Border(top=MEDIUM, bottom=NO_BORDER, left=NO_BORDER, right=MEDIUM)

    # Profitability Story
    ws.merge_cells('H5:I5')
    _set_cell(ws, 5, _col('H'), 'Profitability Story', font=FONT_10_BOLD,
              border=Border(top=MEDIUM, bottom=NO_BORDER, left=MEDIUM, right=MEDIUM))
    ws.cell(row=5, column=_col('I')).border = Border(top=MEDIUM, bottom=NO_BORDER, left=NO_BORDER, right=MEDIUM)

    # Growth Efficiency Story
    ws.merge_cells('K5:L5')
    _set_cell(ws, 5, _col('K'), 'Growth Efficiency Story', font=FONT_10_BOLD,
              border=Border(top=MEDIUM, bottom=NO_BORDER, left=MEDIUM, right=MEDIUM))
    ws.cell(row=5, column=_col('L')).border = Border(top=MEDIUM, bottom=NO_BORDER, left=NO_BORDER, right=MEDIUM)

    # ── Row 6: Sub-headers + Story text boxes ──
    _set_cell(ws, 6, _col('B'), 'Company', font=FONT_10)
    _set_cell(ws, 6, _col('C'), 'Industry', font=FONT_10)

    # Story text in merged cells with background fill
    ws.merge_cells('E6:F12')
    _set_cell(ws, 6, _col('E'), data.growth_story, font=FONT_10, fill=STORY_FILL, alignment=ALIGN_WRAP)
    # Apply fill+border to merged area
    for r in range(6, 13):
        for c in [_col('E'), _col('F')]:
            cell = ws.cell(row=r, column=c)
            cell.fill = STORY_FILL
            cell.border = Border(
                left=MEDIUM if c == _col('E') else NO_BORDER,
                right=MEDIUM if c == _col('F') else NO_BORDER,
                bottom=MEDIUM if r == 12 else NO_BORDER
            )

    ws.merge_cells('H6:I12')
    _set_cell(ws, 6, _col('H'), data.profitability_story, font=FONT_10, fill=STORY_FILL, alignment=ALIGN_WRAP)
    for r in range(6, 13):
        for c in [_col('H'), _col('I')]:
            cell = ws.cell(row=r, column=c)
            cell.fill = STORY_FILL
            cell.border = Border(
                left=MEDIUM if c == _col('H') else NO_BORDER,
                right=MEDIUM if c == _col('I') else NO_BORDER,
                bottom=MEDIUM if r == 12 else NO_BORDER
            )

    ws.merge_cells('K6:L12')
    _set_cell(ws, 6, _col('K'), data.growth_efficiency_story, font=FONT_10, fill=STORY_FILL, alignment=ALIGN_WRAP)
    for r in range(6, 13):
        for c in [_col('K'), _col('L')]:
            cell = ws.cell(row=r, column=c)
            cell.fill = STORY_FILL
            cell.border = Border(
                left=MEDIUM if c == _col('K') else NO_BORDER,
                right=MEDIUM if c == _col('L') else NO_BORDER,
                bottom=MEDIUM if r == 12 else NO_BORDER
            )

    # ── Terminal Value box (O6:P10) ──
    ws.merge_cells('O6:P6')
    _set_cell(ws, 6, _col('O'), 'Terminal Value', font=FONT_10, border=BORDER_ALL_THIN)
    ws.cell(row=6, column=_col('P')).border = Border(top=THIN, right=THIN)

    tv_items = [
        ('Growth Rate', data.tv_growth_rate),
        ('Cost of capital', data.tv_cost_of_capital),
        ('Return on capital', data.tv_return_on_capital),
        ('Reinvestment Rate', data.tv_reinvestment_rate),
    ]
    for i, (label, val) in enumerate(tv_items):
        row = 7 + i
        _set_cell(ws, row, _col('O'), label, font=FONT_10,
                  border=Border(top=THIN, bottom=THIN, left=THIN, right=THIN))
        _set_cell(ws, row, _col('P'), val, font=FONT_10, number_format=NF_PCT,
                  border=Border(top=THIN, bottom=THIN, left=THIN, right=MEDIUM))

    # ── Row 7-11: Base Year data ──
    base_rows = [
        ('Revenue Growth', data.base_revenue_growth, data.industry_revenue_growth, NF_PCT),
        ('Revenue', data.base_revenue, None, nf_dollar),
        ('Operating Margin', data.base_operating_margin, data.industry_operating_margin, NF_PCT),
        ('Operating Income', data.base_operating_income, None, nf_dollar),
        ('EBIT (1-t)', data.base_ebit_1_t, None, nf_dollar),
    ]
    for i, (label, company_val, industry_val, nf) in enumerate(base_rows):
        row = 7 + i
        _set_cell(ws, row, 1, label, font=FONT_10,
                  border=Border(top=THIN, bottom=THIN, left=MEDIUM, right=THIN))
        _set_cell(ws, row, 2, company_val, font=FONT_10, number_format=nf,
                  border=Border(top=THIN, bottom=THIN, left=THIN, right=THIN))
        if industry_val is not None:
            _set_cell(ws, row, 3, industry_val, font=FONT_10, number_format=nf,
                      border=Border(top=THIN, bottom=THIN, left=THIN, right=THIN))
        else:
            ws.cell(row=row, column=3).border = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

    # ── Row 15: Year headers for projection ──
    for i in range(10):
        _set_cell(ws, 15, _col('E') + i, i + 1, font=FONT_10, alignment=ALIGN_CENTER,
                  border=BORDER_ALL_THIN)
    _set_cell(ws, 15, _col('O'), 'Terminal year', font=FONT_10, fill=HIGHLIGHT_FILL,
              border=BORDER_ALL_THIN, alignment=ALIGN_CENTER)

    # ── Left side: Equity Bridge (A15-B30) ──
    bridge_items = [
        (15, 'PV(Terminal value)', data.pv_terminal_value, nf_acct),
        (16, 'PV (CF over next 10 years)', data.pv_cf_next_10, nf_acct),
        (17, 'Probability of failure =', data.probability_of_failure, NF_PCT),
        (18, 'Value of operating assets =', data.value_of_operating_assets, nf_dollar),
        (19, ' - Debt', data.debt, nf_dollar),
        (20, ' - Minority interests', data.minority_interests, nf_dollar),
        (21, ' +  Cash', data.cash, nf_dollar),
        (22, ' + Non-operating assets', data.non_operating_assets, nf_dollar),
        (23, 'Value of equity', data.value_of_equity, nf_dollar),
        (24, ' - Value of options', data.value_of_options, nf_dollar),
        (25, 'Value of equity in common stock', data.value_equity_common, nf_dollar),
        (26, 'Number of shares', data.number_of_shares, NF_SHARES),
        (27, 'Estimated value /share', data.estimated_value_per_share, nf_dollar_dec),
        (29, 'Price per share', data.price_per_share, nf_dollar_dec),
        (30, '% Under or Over Valued', data.pct_under_over_valued, NF_PCT),
    ]
    for row, label, val, nf in bridge_items:
        _set_cell(ws, row, 1, label, font=FONT_10, fill=HIGHLIGHT_FILL)
        b_nf = NF_PCT if row == 17 else nf
        _set_cell(ws, row, 2, val, font=FONT_10, fill=HIGHLIGHT_FILL, number_format=b_nf)

    # ── Right side: 10-year projection table (D16-O29) ──
    proj_label_col = _col('D')
    proj_start_col = _col('E')  # columns E through N = years 1-10, O = terminal

    proj_rows_data = [
        (16, 'Revenue Growth', data.revenue_growth, data.terminal_revenue_growth, NF_PCT, False),
        (17, 'Revenue', data.revenues, data.terminal_revenue, nf_acct, False),
        (18, 'Operating Margin', data.operating_margin, data.terminal_operating_margin, NF_PCT, False),
        (19, 'Operating Income', data.operating_income, data.terminal_operating_income, nf_acct, False),
        (20, 'EBIT (1-t)', data.ebit_1_t, data.terminal_ebit_1_t, nf_acct, False),
        (21, 'Reinvestment', data.reinvestment, data.terminal_reinvestment, nf_acct, False),
        (22, 'FCFF', data.fcff, data.terminal_fcff, nf_acct, True),
        # Skip rows 23-24 (empty in projection area, or terminal value display)
        (25, 'Cost of Capital', data.cost_of_capital, None, NF_PCT, True),
        (26, 'Cumulated WACC', data.cumulated_wacc, None, NF_NUM_4DEC, False),
        (28, 'Sales to Capital', data.sales_to_capital, None, NF_NUM_DEC, True),
        (29, 'ROIC', data.roic, data.terminal_roic, NF_PCT, False),
    ]

    for row, label, values, terminal_val, nf, highlighted in proj_rows_data:
        fill = HIGHLIGHT_FILL if highlighted else None
        _set_cell(ws, row, proj_label_col, label, font=FONT_10, fill=fill,
                  border=BORDER_ALL_THIN)
        for i, v in enumerate(values):
            _set_cell(ws, row, proj_start_col + i, v, font=FONT_10, number_format=nf,
                      border=BORDER_ALL_THIN, fill=fill if highlighted else None)
        if terminal_val is not None:
            _set_cell(ws, row, _col('O'), terminal_val, font=FONT_10, fill=HIGHLIGHT_FILL,
                      number_format=nf, border=BORDER_ALL_THIN)

    # Terminal value total in N23
    _set_cell(ws, 23, _col('N'), data.terminal_value_total, font=FONT_10, fill=HIGHLIGHT_FILL,
              number_format=nf_acct_dec, border=BORDER_ALL_THIN)

    # ── Row 32-38: Risk Story + Competitive Advantages ──
    ws.merge_cells('D32:E32')
    _set_cell(ws, 32, _col('D'), 'Risk Story', font=FONT_10_BOLD)
    ws.merge_cells('G32:I32')
    _set_cell(ws, 32, _col('G'), 'Competitive Advantages', font=FONT_10_BOLD)

    ws.merge_cells('D33:E38')
    _set_cell(ws, 33, _col('D'), data.risk_story, font=FONT_10, fill=STORY_FILL, alignment=ALIGN_WRAP)
    for r in range(33, 39):
        for c in [_col('D'), _col('E')]:
            ws.cell(row=r, column=c).fill = STORY_FILL

    ws.merge_cells('G33:I38')
    _set_cell(ws, 33, _col('G'), data.competitive_advantages, font=FONT_10, fill=STORY_FILL, alignment=ALIGN_WRAP)
    for r in range(33, 39):
        for c in [_col('G'), _col('H'), _col('I')]:
            ws.cell(row=r, column=c).fill = STORY_FILL

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A4

    return ws


# ──────────────────────────────────────────────────────────────────
# Diagnostics Sheet
# ──────────────────────────────────────────────────────────────────

def create_diagnostics(wb, data: DCFInput):
    """Create the Diagnostics sheet — exact Damodaran layout."""
    ws = wb.create_sheet('Diagnostics')

    sym = data.currency_symbol
    nf_dollar_dec = _make_currency_format_dec(sym)

    # Column widths
    ws.column_dimensions['A'].width = 42
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 4
    ws.column_dimensions['G'].width = 80

    # ── Step 1: Revenue Growth Rate ──
    _set_cell(ws, 1, 1, 'Step 1: Check revenue growth rate', font=FONT_12_BOLD)
    ws.merge_cells('D1:E1')
    _set_cell(ws, 1, 4, 'Your forecasts', font=FONT_12_BOLD)
    _set_cell(ws, 1, 7, 'Questions to ask', font=FONT_12_BOLD)

    headers_1 = ['Industry Average', 'Most Recent year', 'Next year', 'Years 2-5']
    for i, h in enumerate(headers_1):
        _set_cell(ws, 2, 2 + i, h, font=FONT_12)

    _set_cell(ws, 2, 7, '1. If you are forecasting a revenue growth rate > industry average, is your company small?', font=FONT_12)
    _set_cell(ws, 3, 1, 'Annual Revenue Growth Rate', font=FONT_12)
    _set_cell(ws, 3, 2, data.industry_avg_revenue_growth, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 3, 3, data.base_revenue_growth, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 3, 4, data.revenue_growth[0] if data.revenue_growth else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 3, 5, data.revenue_growth[1] if len(data.revenue_growth) > 1 else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 3, 7, "2. If your forecasted revenue growth rate is very different from your company's most recent year of growth, what is the reason?", font=FONT_12)

    # ── Step 2: Dollar Revenues ──
    _set_cell(ws, 5, 1, 'Step 2: Check dollar revenues', font=FONT_12_BOLD)
    headers_2 = ['Most Recent year', 'Next year', 'Year 5', 'Year 10']
    for i, h in enumerate(headers_2):
        _set_cell(ws, 6, 2 + i, h, font=FONT_12)
    _set_cell(ws, 6, 7, '1. How big is the total market today?', font=FONT_12)

    _set_cell(ws, 7, 1, 'Revenues', font=FONT_12)
    _set_cell(ws, 7, 2, data.base_revenue, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 7, 3, data.revenues[0] if data.revenues else 0, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 7, 4, data.revenues[4] if len(data.revenues) > 4 else 0, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 7, 5, data.revenues[9] if len(data.revenues) > 9 else 0, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 7, 7, '2. How much revenues do the biggest companies in that market make today?', font=FONT_12)
    _set_cell(ws, 8, 7, '3. How much growth is there in the total market?', font=FONT_12)
    _set_cell(ws, 9, 7, '4. What type of market share are you forecasting for your company in year 10?', font=FONT_12)

    # ── Step 3: Margins ──
    _set_cell(ws, 11, 1, 'Step 3: Check your margins', font=FONT_12_BOLD)
    for i, h in enumerate(headers_2):
        _set_cell(ws, 11, 2 + i, h, font=FONT_12)
    _set_cell(ws, 11, 7, '1. What are the margins of the industry that the company is in?', font=FONT_12)

    _set_cell(ws, 12, 1, 'Operating Margin', font=FONT_12)
    _set_cell(ws, 12, 2, data.base_operating_margin, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 12, 3, data.operating_margin[0] if data.operating_margin else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 12, 4, data.operating_margin[4] if len(data.operating_margin) > 4 else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 12, 5, data.operating_margin[9] if len(data.operating_margin) > 9 else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 12, 7, '2. What are the unit economics of the business? (How much does it cost you to make the extra unit that you sell?', font=FONT_12)
    _set_cell(ws, 13, 7, '3. What does the competition in this business look like?', font=FONT_12)

    # ── Step 4: Reinvestment ──
    _set_cell(ws, 15, 1, 'Step 4: Check how much you are reinvesting', font=FONT_12_BOLD)
    for i, h in enumerate(headers_2):
        _set_cell(ws, 16, 2 + i, h, font=FONT_12)
    _set_cell(ws, 16, 7, '1. is the growth that you are forecasting bounce-back growth or new growth?', font=FONT_12)

    _set_cell(ws, 17, 1, 'Sales to Capital', font=FONT_12)
    _set_cell(ws, 17, 2, data.base_sales_to_capital, font=FONT_12, number_format=NF_NUM_DEC)
    _set_cell(ws, 17, 3, data.sales_to_capital[0] if data.sales_to_capital else 0, font=FONT_12, number_format=NF_NUM_DEC)
    _set_cell(ws, 17, 4, data.sales_to_capital[4] if len(data.sales_to_capital) > 4 else 0, font=FONT_12, number_format=NF_NUM_DEC)
    _set_cell(ws, 17, 5, data.sales_to_capital[9] if len(data.sales_to_capital) > 9 else 0, font=FONT_12, number_format=NF_NUM_DEC)
    _set_cell(ws, 17, 7, '2. How much excess capacity do you have to service near term growth?', font=FONT_12)
    _set_cell(ws, 18, 7, '3. Does investment efficiency in this business change as companies get bigger?', font=FONT_12)

    # Reinvestment effect
    _set_cell(ws, 19, 1, 'Reinvestment effect on cash flows', font=FONT_12)
    _set_cell(ws, 20, 2, '$ Value', font=FONT_12)
    _set_cell(ws, 20, 3, 'As % of value', font=FONT_12)

    _set_cell(ws, 21, 1, 'PV of after-tax operating income for next 10 yearas', font=FONT_12)
    _set_cell(ws, 21, 2, data.pv_ati_10yr, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 21, 3, data.pv_ati_pct, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 21, 7, '1. Is your reinvestment consitent with your revenue growth forecast?', font=FONT_12)

    _set_cell(ws, 22, 1, 'Value effect of reinvestment for next 10 years', font=FONT_12)
    _set_cell(ws, 22, 2, data.reinvestment_value_effect, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 22, 3, data.reinvestment_value_pct, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 22, 7, '2. Are you comfortable with your return on capital in year 10?', font=FONT_12)

    _set_cell(ws, 23, 1, 'PV of FCFF for next ten years', font=FONT_12)
    _set_cell(ws, 23, 2, data.pv_fcff_10yr, font=FONT_12, number_format=nf_dollar_dec)
    _set_cell(ws, 23, 3, data.pv_fcff_pct, font=FONT_12, number_format=NF_PCT)

    # ROC effects
    _set_cell(ws, 25, 1, 'Return on capital effects', font=FONT_12)
    roc_headers = ['Most recent year', 'Marginal (1-10)', 'ROC in year 10', 'Stable ROC']
    for i, h in enumerate(roc_headers):
        _set_cell(ws, 26, 2 + i, h, font=FONT_12)

    _set_cell(ws, 27, 1, 'Return on capital', font=FONT_12)
    _set_cell(ws, 27, 2, data.roc_most_recent, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 27, 3, data.roc_marginal, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 27, 4, data.roc_year10, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 27, 5, data.roc_stable, font=FONT_12, number_format=NF_PCT)

    # ── Step 5: Risk Metrics ──
    _set_cell(ws, 29, 1, 'Step 5: Risk Metrics', font=FONT_12_BOLD)
    risk_headers = ['Industry Average', 'Compounded', 'Year 1-5', 'Stable Growth']
    for i, h in enumerate(risk_headers):
        _set_cell(ws, 30, 2 + i, h, font=FONT_12)
    _set_cell(ws, 30, 7, '1. How does your cost of capital compare to the industry average?', font=FONT_12)

    _set_cell(ws, 31, 1, 'Cost of capital', font=FONT_12)
    _set_cell(ws, 31, 2, data.industry_avg_cost_of_capital, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 31, 3, data.compounded_cost_of_capital, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 31, 4, data.cost_of_capital[0] if data.cost_of_capital else 0, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 31, 5, data.terminal_cost_of_capital, font=FONT_12, number_format=NF_PCT)
    _set_cell(ws, 31, 7, '2. What is happenign to your cost of capital over time? Why?', font=FONT_12)
    _set_cell(ws, 32, 7, "3. Is your failure rate consistent with your company's characteristics?", font=FONT_12)

    _set_cell(ws, 33, 1, 'Failure Rate', font=FONT_12)
    _set_cell(ws, 33, 2, data.failure_rate, font=FONT_12, number_format=NF_PCT)

    # ── Step 6: Price vs Value ──
    _set_cell(ws, 35, 1, 'Step 6: Price versus Value', font=FONT_12_BOLD)
    _set_cell(ws, 36, 1, 'Your calculated value as a percent of current price', font=Font(name='Calibri', size=10))
    _set_cell(ws, 36, 2, data.value_as_pct_of_price, font=Font(name='Calibri', size=10), number_format=NF_PCT)

    # Suggestion text
    ws.merge_cells('B37:C37')
    suggestion = 'Value seems low. See below' if data.value_as_pct_of_price < 0.5 else \
                 'Value seems high. See below' if data.value_as_pct_of_price > 2.0 else ''
    if suggestion:
        _set_cell(ws, 37, 2, suggestion, font=FONT_12)

    # Adjustment table
    _set_cell(ws, 38, 1, 'Inputs', font=Font(name='Calibri', size=10, bold=True))
    _set_cell(ws, 38, 2, 'If calculated value is negative or looks too low', font=Font(name='Calibri', size=10, bold=True))
    _set_cell(ws, 38, 3, 'If calculated value looks too high', font=Font(name='Calibri', size=10, bold=True))

    adjustments = [
        ('Revenue growth rate (input cell B25, B27)', 'Increase revenue growth rate', 'Decrease revenue growth rate'),
        ('Operating margin (B26, B28)', 'Increase the target pre-tax operating margin', 'Decrease the target pre-tax operating margin'),
        ('Sales to Capital (B30-B32)', 'Increase the sales/capital ratio', 'Decrease the sales/capital ratio'),
        ('Return on capital in perpetuity (B48, B49)', 'Increase relative to your cost of capital', 'If higher than your cost of capital, lower towards your cost of capital'),
    ]
    f10 = Font(name='Calibri', size=10)
    for i, (inp, low, high) in enumerate(adjustments):
        r = 39 + i
        _set_cell(ws, r, 1, inp, font=f10)
        _set_cell(ws, r, 2, low, font=f10)
        _set_cell(ws, r, 3, high, font=f10)

    # Print settings
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'

    return ws


# ──────────────────────────────────────────────────────────────────
# Main Generator
# ──────────────────────────────────────────────────────────────────

def generate_dcf_output(data: DCFInput, output_dir: str, filename_base: str = "dcf_output"):
    """Generate Excel (and optionally PDF) DCF output files."""
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    create_valuation_picture(wb, data)
    create_diagnostics(wb, data)

    # Save Excel
    excel_path = os.path.join(output_dir, f"{filename_base}.xlsx")
    wb.save(excel_path)
    print(f"Excel saved: {excel_path}")

    # Try PDF generation via LibreOffice if available
    pdf_path = os.path.join(output_dir, f"{filename_base}.pdf")
    try:
        import subprocess
        result = subprocess.run(
            ['libreoffice', '--headless', '--calc', '--convert-to', 'pdf',
             '--outdir', output_dir, excel_path],
            capture_output=True, text=True, timeout=60
        )
        if result.returncode == 0:
            print(f"PDF saved: {pdf_path}")
        else:
            print(f"PDF generation failed: {result.stderr}")
            # Try soffice as fallback
            result2 = subprocess.run(
                ['soffice', '--headless', '--calc', '--convert-to', 'pdf',
                 '--outdir', output_dir, excel_path],
                capture_output=True, text=True, timeout=60
            )
            if result2.returncode == 0:
                print(f"PDF saved: {pdf_path}")
            else:
                print("PDF generation not available (LibreOffice not found)")
    except (FileNotFoundError, subprocess.TimeoutExpired):
        print("PDF generation skipped (LibreOffice not installed)")

    return excel_path


# ──────────────────────────────────────────────────────────────────
# Test with dummy data
# ──────────────────────────────────────────────────────────────────

def test_with_dummy_data():
    """Generate test output with Almarai-like dummy data."""
    data = DCFInput(
        company_name="Test Company (Dummy Data)",
        valuation_date=datetime(2026, 2, 1),
        currency_symbol="$",

        # Base year
        base_revenue_growth=0.0763,
        industry_revenue_growth=0.0965,
        base_revenue=21765,
        base_operating_margin=0.1406,
        industry_operating_margin=0.0854,
        base_operating_income=3061,
        base_ebit_1_t=2525,

        # Stories
        growth_story="Revenue growth driven by domestic market expansion and new product lines in dairy and poultry.",
        profitability_story="Margins expected to remain stable with gradual improvement from operational efficiencies.",
        growth_efficiency_story="Reinvestment maintained at current levels to support capacity expansion.",
        risk_story="Cost of capital reflects emerging market exposure. Low probability of failure given market dominance.",
        competitive_advantages="Dominant market share in GCC dairy. Vertically integrated supply chain. Strong brand loyalty.",

        # Projections
        revenue_growth=[0.05, 0.05, 0.05, 0.05, 0.05, 0.049, 0.048, 0.047, 0.047, 0.046],
        revenues=[22854, 23996, 25196, 26456, 27779, 29144, 30553, 32003, 33496, 35030],
        operating_margin=[0.1406]*10,
        operating_income=[3214, 3375, 3543, 3721, 3907, 4099, 4297, 4501, 4711, 4926],
        ebit_1_t=[2652, 2784, 2923, 3069, 3223, 3320, 3416, 3511, 3604, 3695],
        reinvestment=[669, 702, 737, 774, 799, 824, 849, 874, 898, 939],
        fcff=[1983, 2082, 2186, 2295, 2424, 2496, 2567, 2637, 2706, 2756],
        cost_of_capital=[0.0706, 0.0706, 0.0706, 0.0706, 0.0706, 0.0741, 0.0776, 0.0811, 0.0846, 0.0881],
        cumulated_wacc=[0.934, 0.873, 0.815, 0.761, 0.711, 0.662, 0.614, 0.568, 0.524, 0.482],
        sales_to_capital=[1.71]*10,
        roic=[0.072, 0.074, 0.077, 0.079, 0.081, 0.082, 0.083, 0.083, 0.084, 0.084],

        # Terminal
        terminal_revenue_growth=0.0458,
        terminal_revenue=36634,
        terminal_operating_margin=0.1406,
        terminal_operating_income=5152,
        terminal_ebit_1_t=3864,
        terminal_reinvestment=2009,
        terminal_fcff=1855,
        terminal_cost_of_capital=0.0881,
        terminal_roic=0.0881,

        # TV box
        tv_growth_rate=0.0458,
        tv_cost_of_capital=0.0881,
        tv_return_on_capital=0.0881,
        tv_reinvestment_rate=0.52,

        # Equity bridge
        pv_terminal_value=21123,
        pv_cf_next_10=16395,
        probability_of_failure=0.0,
        value_of_operating_assets=37518,
        debt=45063,
        minority_interests=1558,
        cash=19000,
        non_operating_assets=21119,
        value_of_equity=31016,
        value_of_options=0,
        value_equity_common=31016,
        number_of_shares=4315,
        estimated_value_per_share=7.19,
        price_per_share=72.28,
        pct_under_over_valued=9.06,
        terminal_value_total=43859,

        # Diagnostics
        industry_avg_cost_of_capital=0.0579,
        compounded_cost_of_capital=0.0758,
        failure_rate=0.0,
        industry_avg_revenue_growth=0.0716,
        base_sales_to_capital=0.59,
        pv_ati_10yr=21878,
        pv_ati_pct=1.0,
        reinvestment_value_effect=5483,
        reinvestment_value_pct=0.2506,
        pv_fcff_10yr=16395,
        pv_fcff_pct=0.7494,
        roc_most_recent=0.0687,
        roc_marginal=0.145,
        roc_year10=0.0842,
        roc_stable=0.0881,
        value_as_pct_of_price=0.0994,
    )

    output_dir = os.path.expanduser("~/.openclaw/workspace/temp")
    excel_path = generate_dcf_output(data, output_dir, "dcf_test_output")
    return excel_path


if __name__ == '__main__':
    test_with_dummy_data()
