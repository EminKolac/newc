#!/usr/bin/env python3
"""
DCF Excel Generator - BBB Toolkit v2.0
Damodaran FCFF + Turkiye Metodolojisi + Profesyonel Excel Cikti

6 Sheet:
  1. OZET        - Valuation as Picture (Damodaran layout) + Equity Bridge
  2. DCF         - 10Y projeksiyon tablosu (CANLI FORMULLER)
  3. WACC        - Tum WACC bilesenleri
  4. SENSITIVITY - 3 tablo x 5x5 grid (75 hucre, formul bazli)
  5. SCENARIOS   - Bear/Base/Bull varsayim bloklari + case selector
  6. DIAGNOSTICS - 17 Damodaran sorusu

KURAL: Her projeksiyon/hesaplama hucresi CANLI FORMUL.
Hardcoded = sadece ham girdiler (mavi font).
Formul = siyah font. Sheet referansi = yesil font.

Kullanim:
    from dcf_excel_generator import generate_dcf_excel
    from dcf_engine import DCFEngine, DCFResult
    from wacc_calculator import WACCResult

    engine = DCFEngine(...)
    dcf_result = engine.calculate()
    wacc_result = calculate_wacc(...)

    generate_dcf_excel(dcf_result, wacc_result, params, output_path)
"""

import os
import sys
from datetime import datetime
from typing import Optional, List, Dict, Any, Tuple
from copy import copy

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Border, Side, Alignment,
        numbers, NamedStyle
    )
    from openpyxl.utils import get_column_letter, column_index_from_string
    from openpyxl.chart import LineChart, Reference
except ImportError:
    print("openpyxl gerekli: pip install openpyxl")
    sys.exit(1)


# ============================================================
# RENKLER & STILLER
# ============================================================

# --- Fills ---
FILL_DARK_HEADER = PatternFill(patternType='solid', fgColor='1F4E79')   # Koyu mavi header
FILL_SUB_HEADER  = PatternFill(patternType='solid', fgColor='D9E1F2')   # Acik mavi sub-header
FILL_INPUT       = PatternFill(patternType='solid', fgColor='F2F2F2')   # Acik gri input
FILL_OUTPUT      = PatternFill(patternType='solid', fgColor='BDD7EE')   # Orta mavi output
FILL_STORY       = PatternFill(patternType='solid', fgColor='DAEEF3')   # Hikaye kutusu
FILL_WARN        = PatternFill(patternType='solid', fgColor='FFF2CC')   # Sari uyari
FILL_GREEN_LIGHT = PatternFill(patternType='solid', fgColor='E2EFDA')   # Yesil vurgu
FILL_TERMINAL    = PatternFill(patternType='solid', fgColor='E2EFDA')   # Terminal yesil
FILL_WHITE       = PatternFill(patternType='solid', fgColor='FFFFFF')
FILL_NONE        = PatternFill(fill_type=None)

# --- Fonts ---
FONT_HEADER_W = Font(name='Calibri', size=11, bold=True, color='FFFFFF')  # Beyaz bold
FONT_SUBHEADER = Font(name='Calibri', size=10, bold=True, color='000000')  # Siyah bold
FONT_INPUT     = Font(name='Calibri', size=10, color='0000FF')             # Mavi (input)
FONT_FORMULA   = Font(name='Calibri', size=10, color='000000')             # Siyah (formul)
FONT_LINK      = Font(name='Calibri', size=10, color='008000')             # Yesil (sheet link)
FONT_TITLE     = Font(name='Calibri', size=14, bold=True, color='1F4E79')
FONT_SUBTITLE  = Font(name='Calibri', size=11, bold=True, color='1F4E79')
FONT_OUTPUT    = Font(name='Calibri', size=11, bold=True, color='000000')
FONT_WARN      = Font(name='Calibri', size=10, color='C00000')
FONT_SMALL     = Font(name='Calibri', size=9, color='808080')
FONT_10        = Font(name='Calibri', size=10)
FONT_10B       = Font(name='Calibri', size=10, bold=True)

# --- Borders ---
THIN   = Side(style='thin')
MEDIUM = Side(style='medium')
THICK  = Side(style='thick')
HAIR   = Side(style='hair')
NONE   = Side(style=None)

BORDER_THIN     = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_MEDIUM   = Border(top=MEDIUM, bottom=MEDIUM, left=MEDIUM, right=MEDIUM)
BORDER_BOTTOM_M = Border(bottom=MEDIUM)
BORDER_TOP_M    = Border(top=MEDIUM)
BORDER_BOTTOM_T = Border(bottom=THIN)
BORDER_HAIR     = Border(top=HAIR, bottom=HAIR, left=HAIR, right=HAIR)
BORDER_SECTION  = Border(top=THICK, bottom=THICK, left=THICK, right=THICK)

# --- Alignment ---
ALIGN_L   = Alignment(horizontal='left', vertical='center')
ALIGN_C   = Alignment(horizontal='center', vertical='center')
ALIGN_R   = Alignment(horizontal='right', vertical='center')
ALIGN_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)

# --- Number Formats ---
NF_PCT     = '0.0%'
NF_PCT2    = '0.00%'
NF_NUM     = '#,##0'
NF_NUM1    = '#,##0.0'
NF_NUM2    = '#,##0.00'
NF_TL      = '₺#,##0'
NF_TL2     = '₺#,##0.00'
NF_USD     = '$#,##0'
NF_USD2    = '$#,##0.00'
NF_X       = '0.0"x"'
NF_DATE    = 'DD.MM.YYYY'


# ============================================================
# HELPER FONKSIYONLAR
# ============================================================

def _cell(ws, row, col, value=None, font=None, fill=None, border=None,
          alignment=None, nf=None, comment_text=None):
    """Hucre yaz + formatla."""
    c = ws.cell(row=row, column=col, value=value)
    if font: c.font = font
    if fill: c.fill = fill
    if border: c.border = border
    if alignment: c.alignment = alignment
    if nf: c.number_format = nf
    if comment_text:
        from openpyxl.comments import Comment
        c.comment = Comment(comment_text, "BBB DCF v2.0")
    return c


def _merge_header(ws, row, col_start, col_end, text, font=FONT_HEADER_W,
                  fill=FILL_DARK_HEADER, alignment=ALIGN_L, border=None):
    """Merge + header yaz."""
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row, end_column=col_end)
    _cell(ws, row, col_start, text, font=font, fill=fill, alignment=alignment, border=border)
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
        if border:
            ws.cell(row=row, column=c).border = border


def _apply_border_box(ws, r1, c1, r2, c2, border_style=MEDIUM):
    """Bolgeye dis kenar ciz."""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            new_border = Border(
                top=border_style if r == r1 else (cell.border.top or NONE),
                bottom=border_style if r == r2 else (cell.border.bottom or NONE),
                left=border_style if c == c1 else (cell.border.left or NONE),
                right=border_style if c == c2 else (cell.border.right or NONE),
            )
            cell.border = new_border


def _col_letter(n):
    return get_column_letter(n)


def _nf_currency(currency):
    if currency == "TL":
        return NF_TL, NF_TL2
    return NF_USD, NF_USD2


# ============================================================
# SHEET 1: OZET (Valuation as Picture)
# ============================================================

def _build_ozet(wb, dcf, wacc, params):
    """Damodaran 'Valuation as Picture' + Equity Bridge."""
    ws = wb.create_sheet("OZET")
    cur = params.get("currency", "TL")
    nf_c, nf_c2 = _nf_currency(cur)
    ticker = params.get("ticker", "XXXX")
    company = params.get("company_name", ticker)
    date_str = params.get("date", datetime.now().strftime("%d.%m.%Y"))

    # Sutun genislikleri
    widths = [30, 14, 14, 4, 18, 14, 14, 14, 14, 14, 14, 14, 14, 14, 18, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[_col_letter(i)].width = w

    # === ROW 1-2: BASLIK ===
    ws.merge_cells('A1:P1')
    _cell(ws, 1, 1, f"{company} ({ticker}) - DCF Degerleme", font=FONT_TITLE)
    ws.merge_cells('A2:H2')
    _cell(ws, 2, 1, f"Tarih: {date_str} | Para Birimi: {cur} (Milyon) | Damodaran FCFF Metodolojisi",
          font=FONT_SMALL)

    # === ROW 4: BAZ YIL ===
    _merge_header(ws, 4, 1, 3, "BAZ YIL VERiLERi")
    base_year = params.get("base_year", "FY2025")
    items = [
        ("Baz Donem", base_year, None),
        ("Hasilat", dcf.base_revenue, nf_c),
        ("EBIT (Duzeltilmis)", dcf.base_ebit, nf_c),
        ("EBIT Marji", dcf.base_ebit_margin / 100, NF_PCT),
    ]
    for i, (label, val, nf) in enumerate(items):
        r = 5 + i
        _cell(ws, r, 1, label, font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=nf)
    _apply_border_box(ws, 4, 1, 8, 3)

    # === ROW 4-12: HIKAYE KUTULARI ===
    stories = [
        (5, "Buyume Hikayesi", params.get("growth_story", "")),
        (8, "Karlilik Hikayesi", params.get("profitability_story", "")),
        (11, "Rekabet Avantaji", params.get("competitive_story", "")),
    ]
    for col_start, title, text in stories:
        _merge_header(ws, 4, col_start, col_start + 1, title)
        ws.merge_cells(start_row=5, start_column=col_start,
                       end_row=8, end_column=col_start + 1)
        _cell(ws, 5, col_start, text, font=FONT_10, fill=FILL_STORY, alignment=ALIGN_WRAP)
        for r in range(5, 9):
            for c in range(col_start, col_start + 2):
                ws.cell(row=r, column=c).fill = FILL_STORY

    # === ROW 4: TERMINAL KUTUSU ===
    _merge_header(ws, 4, 15, 16, "TERMINAL DEGER")
    t = dcf.terminal
    tv_items = [
        ("Terminal Buyume", t.terminal_growth / 100, NF_PCT),
        ("Terminal WACC", t.terminal_wacc / 100, NF_PCT),
        ("Terminal ROC", t.terminal_roc / 100, NF_PCT),
        ("Reinvestment Rate", t.terminal_reinvestment_rate / 100, NF_PCT),
    ]
    for i, (label, val, nf) in enumerate(tv_items):
        r = 5 + i
        _cell(ws, r, 15, label, font=FONT_10, fill=FILL_TERMINAL, border=BORDER_THIN)
        _cell(ws, r, 16, val, font=FONT_INPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf)
    _apply_border_box(ws, 4, 15, 8, 16)

    # === ROW 10: PROJEKSIYON TABLOSU ===
    _merge_header(ws, 10, 1, 16, "10 YILLIK PROJEKSiYON")

    # Yil basliklari
    _cell(ws, 11, 1, "", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
    _cell(ws, 11, 2, "Baz", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN, alignment=ALIGN_C)
    for y in range(10):
        _cell(ws, 11, 3 + y, f"Y{y+1}", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)
    _cell(ws, 11, 13, "Terminal", font=FONT_10B, fill=FILL_TERMINAL, border=BORDER_THIN, alignment=ALIGN_C)

    # Veri satirlari
    proj_rows = [
        ("Gelir Buyumesi", None, [p.revenue_growth / 100 for p in dcf.projections],
         t.terminal_growth / 100, NF_PCT),
        ("Hasilat", dcf.base_revenue, [p.revenue for p in dcf.projections],
         t.terminal_revenue, nf_c),
        ("EBIT Marji", dcf.base_ebit_margin / 100, [p.ebit_margin / 100 for p in dcf.projections],
         t.terminal_ebit_margin / 100, NF_PCT),
        ("EBIT", dcf.base_ebit, [p.ebit for p in dcf.projections],
         t.terminal_ebit, nf_c),
        ("EBIT(1-t)", None, [p.ebit_1_t for p in dcf.projections],
         t.terminal_ebit_1_t, nf_c),
        ("Reinvestment", None, [p.reinvestment for p in dcf.projections],
         None, nf_c),
        ("FCFF", None, [p.fcff for p in dcf.projections],
         t.terminal_fcff, nf_c),
        ("", None, [None]*10, None, None),
        ("WACC", None, [p.wacc / 100 for p in dcf.projections],
         t.terminal_wacc / 100, NF_PCT),
        ("Iskonto Faktoru", None, [p.discount_factor for p in dcf.projections],
         None, '0.0000'),
        ("PV(FCFF)", None, [p.pv_fcff for p in dcf.projections],
         None, nf_c),
    ]

    for i, (label, base_val, values, term_val, nf) in enumerate(proj_rows):
        r = 12 + i
        is_output = label in ("FCFF", "PV(FCFF)")
        fill = FILL_OUTPUT if is_output else FILL_WHITE
        font = FONT_OUTPUT if is_output else FONT_FORMULA

        _cell(ws, r, 1, label, font=FONT_10B if label else FONT_10,
              fill=fill, border=BORDER_THIN)
        if base_val is not None:
            _cell(ws, r, 2, base_val, font=FONT_INPUT, fill=FILL_INPUT,
                  border=BORDER_THIN, nf=nf)

        for y in range(10):
            val = values[y] if y < len(values) else None
            if val is not None:
                _cell(ws, r, 3 + y, val, font=font, fill=fill,
                      border=BORDER_THIN, nf=nf)

        if term_val is not None:
            _cell(ws, r, 13, term_val, font=font, fill=FILL_TERMINAL,
                  border=BORDER_THIN, nf=nf)

    _apply_border_box(ws, 10, 1, 22, 13)

    # === ROW 24: EQUITY BRIDGE ===
    _merge_header(ws, 24, 1, 5, "DEGERLEME KOPRUSU (Equity Bridge)")
    b = dcf.bridge
    bridge_items = [
        ("PV (10Y FCFF)", b.pv_projection, nf_c, False),
        ("PV (Terminal Deger)", b.pv_terminal, nf_c, False),
        ("= Faaliyet Varlik Degeri", b.operating_assets, nf_c, True),
        ("(-) Borc", b.minus_debt, nf_c, False),
        ("(-) Azinlik Paylari", b.minus_minority, nf_c, False),
        ("(+) Nakit ve Benzerleri", b.plus_cash, nf_c, False),
        ("(+) Istirakler", b.plus_cross_holdings, nf_c, False),
        ("(-) Calisan Opsiyonlari", b.minus_options, nf_c, False),
        ("= OZSERMAYE DEGERi", b.equity_value, nf_c, True),
        ("/ Pay Sayisi (M)", b.shares, NF_NUM1, False),
        ("= HiSSE BASINA DEGER", b.value_per_share, nf_c2, True),
    ]

    current_price = params.get("current_price")
    if current_price:
        upside = (b.value_per_share / current_price - 1)
        bridge_items.extend([
            ("", None, None, False),
            ("Mevcut Fiyat", current_price, nf_c2, False),
            ("Potansiyel", upside, NF_PCT, True),
        ])

    for i, (label, val, nf, is_key) in enumerate(bridge_items):
        r = 25 + i
        fill_r = FILL_OUTPUT if is_key else FILL_WHITE
        font_r = FONT_OUTPUT if is_key else FONT_FORMULA
        _cell(ws, r, 1, label, font=font_r, fill=fill_r, border=BORDER_THIN)
        if val is not None:
            _cell(ws, r, 2, val, font=font_r, fill=fill_r, border=BORDER_THIN, nf=nf)

    # TV/EV orani
    _merge_header(ws, 24, 7, 10, "SANiTY CHECK")
    sanity = [
        ("TV / EV", dcf.tv_pct_of_ev / 100, NF_PCT),
    ]
    if dcf.implied_ev_ebitda:
        sanity.append(("Implied EV/EBITDA", dcf.implied_ev_ebitda, NF_X))
    if dcf.implied_pe:
        sanity.append(("Implied P/E", dcf.implied_pe, NF_X))
    sanity.append(("Terminal Buyume", t.terminal_growth / 100, NF_PCT))
    sanity.append(("Terminal WACC", t.terminal_wacc / 100, NF_PCT))

    for i, (label, val, nf) in enumerate(sanity):
        r = 25 + i
        _cell(ws, r, 7, label, font=FONT_10, fill=FILL_SUB_HEADER, border=BORDER_THIN)
        _cell(ws, r, 8, val, font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN, nf=nf)

    # Uyarilar
    if dcf.warnings:
        warn_row = 25 + len(sanity) + 1
        _cell(ws, warn_row, 7, "UYARILAR", font=FONT_WARN, fill=FILL_WARN)
        for j, w in enumerate(dcf.warnings):
            _cell(ws, warn_row + 1 + j, 7, f"! {w}", font=FONT_WARN)

    # Print ayarlari
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.page_setup.orientation = 'landscape'

    return ws


# ============================================================
# SHEET 2: DCF (Formul Bazli Projeksiyon)
# ============================================================

def _build_dcf(wb, dcf, wacc, params):
    """
    Ana DCF sheet'i - CANLI FORMULLER.

    Row mapping:
      3:  Baz yil girdileri
      7:  Senaryo secici (=SCENARIOS!B2)
      9:  Buyume oranlari (=SCENARIOS referansi)
      10: Hasilat (formul: onceki * (1 + buyume))
      11: EBIT marji
      12: EBIT
      13: Vergi orani
      14: EBIT(1-t)
      15: Reinvestment (dRev / S/C)
      16: FCFF
      18: WACC
      19: Kumulatif iskonto
      20: PV(FCFF)
      22: Terminal deger
    """
    ws = wb.create_sheet("DCF")
    cur = params.get("currency", "TL")
    nf_c, nf_c2 = _nf_currency(cur)
    n_years = len(dcf.projections)
    ticker = params.get("ticker", "XXXX")

    # Sutun genislikleri: A=label, B=baz, C-L=Y1-10, M=Terminal
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 16
    for i in range(n_years):
        ws.column_dimensions[_col_letter(3 + i)].width = 14
    ws.column_dimensions[_col_letter(3 + n_years)].width = 16

    term_col = 3 + n_years  # M = terminal

    # === HEADER ===
    _merge_header(ws, 1, 1, term_col, f"{ticker} - FCFF DCF MODEL ({cur})")
    _cell(ws, 2, 1, f"Baz Donem: {params.get('base_year', 'FY2025')} | "
                     f"Tarih: {params.get('date', '')} | Damodaran Metodolojisi",
          font=FONT_SMALL)

    # === BAZ YIL GIRDILERI (hardcoded, mavi) ===
    _merge_header(ws, 3, 1, 5, "BAZ YIL GiRDiLERi")
    base_inputs = [
        ("Hasilat", dcf.base_revenue, nf_c, f"Kaynak: Is Yatirim IAS 29, {params.get('base_year','')}"),
        ("EBIT (Duzeltilmis)", dcf.base_ebit, nf_c, "Duzeltilmis EBIT: vadeli alim add-back + tek seferlik"),
        ("Sales/Capital", params.get("sales_to_capital", 2.0), NF_NUM2, "Kaynak: Tarihsel ortalama"),
        ("Vergi Orani", params.get("tax_rate", 25) / 100, NF_PCT, "Kaynak: KAP, efektif oran"),
    ]
    for i, (label, val, nf, cmt) in enumerate(base_inputs):
        r = 4 + i
        _cell(ws, r, 1, label, font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN,
              nf=nf, comment_text=cmt)

    # === SENARYO SECICI ===
    _cell(ws, 8, 1, "Aktif Senaryo", font=FONT_10B, fill=FILL_WARN)
    _cell(ws, 8, 2, dcf.scenario_name, font=FONT_OUTPUT, fill=FILL_WARN)

    # === PROJEKSiYON TABLOSU ===
    _merge_header(ws, 9, 1, term_col, "10 YILLIK PROJEKSiYON + TERMiNAL")

    # Yil basliklari
    _cell(ws, 10, 1, "", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
    _cell(ws, 10, 2, "Baz", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN, alignment=ALIGN_C)
    for y in range(n_years):
        _cell(ws, 10, 3 + y, f"Y{y+1}", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)
    _cell(ws, 10, term_col, "Terminal", font=FONT_10B, fill=FILL_TERMINAL,
          border=BORDER_THIN, alignment=ALIGN_C)

    # ROW 11: Gelir Buyumesi (INPUT - mavi)
    r = 11
    _cell(ws, r, 1, "Gelir Buyumesi", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, "", border=BORDER_THIN)
    for y in range(n_years):
        _cell(ws, r, 3 + y, dcf.projections[y].revenue_growth / 100,
              font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=NF_PCT,
              comment_text=f"Y{y+1} buyume orani")
    _cell(ws, r, term_col, dcf.terminal.terminal_growth / 100,
          font=FONT_INPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=NF_PCT)

    # ROW 12: Hasilat (FORMUL)
    r = 12
    _cell(ws, r, 1, "Hasilat", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, dcf.base_revenue, font=FONT_INPUT, fill=FILL_INPUT,
          border=BORDER_THIN, nf=nf_c)
    for y in range(n_years):
        col = 3 + y
        prev_col_letter = _col_letter(col - 1)
        growth_cell = f"{_col_letter(col)}11"
        formula = f"={prev_col_letter}12*(1+{growth_cell})"
        _cell(ws, r, col, formula, font=FONT_FORMULA, border=BORDER_THIN, nf=nf_c)
    # Terminal
    prev_letter = _col_letter(term_col - 1)
    tg_cell = f"{_col_letter(term_col)}11"
    _cell(ws, r, term_col, f"={prev_letter}12*(1+{tg_cell})",
          font=FONT_FORMULA, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf_c)

    # ROW 13: EBIT Marji (INPUT)
    r = 13
    _cell(ws, r, 1, "EBIT Marji", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, dcf.base_ebit_margin / 100, font=FONT_INPUT, fill=FILL_INPUT,
          border=BORDER_THIN, nf=NF_PCT)
    for y in range(n_years):
        _cell(ws, r, 3 + y, dcf.projections[y].ebit_margin / 100,
              font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=NF_PCT)
    _cell(ws, r, term_col, dcf.terminal.terminal_ebit_margin / 100,
          font=FONT_INPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=NF_PCT)

    # ROW 14: EBIT (FORMUL = Hasilat x Marj)
    r = 14
    _cell(ws, r, 1, "EBIT", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, dcf.base_ebit, font=FONT_INPUT, fill=FILL_INPUT,
          border=BORDER_THIN, nf=nf_c)
    for y in range(n_years):
        col = 3 + y
        cl = _col_letter(col)
        _cell(ws, r, col, f"={cl}12*{cl}13", font=FONT_FORMULA, border=BORDER_THIN, nf=nf_c)
    cl_t = _col_letter(term_col)
    _cell(ws, r, term_col, f"={cl_t}12*{cl_t}13",
          font=FONT_FORMULA, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf_c)

    # ROW 15: Vergi Orani (INPUT)
    r = 15
    _cell(ws, r, 1, "Vergi Orani", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, params.get("tax_rate", 25) / 100, font=FONT_INPUT, fill=FILL_INPUT,
          border=BORDER_THIN, nf=NF_PCT)
    for y in range(n_years):
        _cell(ws, r, 3 + y, dcf.projections[y].tax_rate / 100,
              font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=NF_PCT)
    _cell(ws, r, term_col, dcf.terminal.terminal_ebit_1_t / dcf.terminal.terminal_ebit
          if dcf.terminal.terminal_ebit != 0 else 0.25,
          font=FONT_INPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=NF_PCT)

    # ROW 16: EBIT(1-t) (FORMUL)
    r = 16
    _cell(ws, r, 1, "EBIT(1-t)", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, "", border=BORDER_THIN)
    for y in range(n_years):
        col = 3 + y
        cl = _col_letter(col)
        _cell(ws, r, col, f"={cl}14*(1-{cl}15)", font=FONT_FORMULA, border=BORDER_THIN, nf=nf_c)
    _cell(ws, r, term_col, f"={cl_t}14*(1-{cl_t}15)",
          font=FONT_FORMULA, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf_c)

    # ROW 17: Sales/Capital (INPUT)
    r = 17
    _cell(ws, r, 1, "Sales/Capital", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, params.get("sales_to_capital", 2.0),
          font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=NF_NUM2)
    for y in range(n_years):
        _cell(ws, r, 3 + y, dcf.projections[y].sales_to_capital,
              font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=NF_NUM2)

    # ROW 18: Reinvestment (FORMUL = dRev / S/C)
    r = 18
    _cell(ws, r, 1, "Reinvestment", font=FONT_10B, fill=FILL_WHITE, border=BORDER_THIN)
    _cell(ws, r, 2, "", border=BORDER_THIN)
    for y in range(n_years):
        col = 3 + y
        cl = _col_letter(col)
        prev = _col_letter(col - 1)
        _cell(ws, r, col, f"=({cl}12-{prev}12)/{cl}17",
              font=FONT_FORMULA, border=BORDER_THIN, nf=nf_c)

    # ROW 19: FCFF (FORMUL = EBIT(1-t) - Reinvestment)
    r = 19
    _cell(ws, r, 1, "FCFF", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, r, 2, "", fill=FILL_OUTPUT, border=BORDER_THIN)
    for y in range(n_years):
        col = 3 + y
        cl = _col_letter(col)
        _cell(ws, r, col, f"={cl}16-{cl}18",
              font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN, nf=nf_c)

    # Terminal FCFF (with ROC-based reinvestment rate)
    # TV_FCFF = EBIT(1-t) * (1 - g/ROC)
    t_roc = dcf.terminal.terminal_roc / 100 if dcf.terminal.terminal_roc > 0 else 0.10
    _cell(ws, r, term_col,
          dcf.terminal.terminal_fcff,
          font=FONT_OUTPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf_c,
          comment_text=f"FCFF = EBIT(1-t) x (1 - g/ROC), ROC={dcf.terminal.terminal_roc}%")

    # ROW 20: bos
    # ROW 21: WACC (INPUT)
    r = 21
    _cell(ws, r, 1, "WACC", font=FONT_10B, fill=FILL_GREEN_LIGHT, border=BORDER_THIN)
    _cell(ws, r, 2, "", fill=FILL_GREEN_LIGHT, border=BORDER_THIN)
    for y in range(n_years):
        _cell(ws, r, 3 + y, dcf.projections[y].wacc / 100,
              font=FONT_LINK, fill=FILL_GREEN_LIGHT, border=BORDER_THIN, nf=NF_PCT,
              comment_text="Kaynak: WACC sheet")
    _cell(ws, r, term_col, dcf.terminal.terminal_wacc / 100,
          font=FONT_LINK, fill=FILL_TERMINAL, border=BORDER_THIN, nf=NF_PCT)

    # ROW 22: Kumulatif Iskonto Faktoru (FORMUL)
    r = 22
    _cell(ws, r, 1, "Kum. iskonto Faktoru", font=FONT_10, border=BORDER_THIN)
    _cell(ws, r, 2, "", border=BORDER_THIN)
    # Y1: = 1 / (1 + WACC_Y1)
    _cell(ws, r, 3, f"=1/(1+C21)", font=FONT_FORMULA, border=BORDER_THIN, nf='0.000000')
    # Y2-10: = onceki / (1 + WACC_Yi)
    for y in range(1, n_years):
        col = 3 + y
        cl = _col_letter(col)
        prev = _col_letter(col - 1)
        _cell(ws, r, col, f"={prev}22/(1+{cl}21)",
              font=FONT_FORMULA, border=BORDER_THIN, nf='0.000000')

    # ROW 23: PV(FCFF) (FORMUL = FCFF x Iskonto)
    r = 23
    _cell(ws, r, 1, "PV(FCFF)", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, r, 2, "", fill=FILL_OUTPUT, border=BORDER_THIN)
    for y in range(n_years):
        col = 3 + y
        cl = _col_letter(col)
        _cell(ws, r, col, f"={cl}19*{cl}22",
              font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN, nf=nf_c)

    # === ROW 25: TERMINAL DEGER HESABI ===
    _merge_header(ws, 25, 1, 6, "TERMiNAL DEGER")
    tv_items = [
        ("Terminal FCFF", dcf.terminal.terminal_fcff, nf_c),
        ("Terminal WACC", dcf.terminal.terminal_wacc / 100, NF_PCT),
        ("Terminal Buyume", dcf.terminal.terminal_growth / 100, NF_PCT),
        ("Terminal ROC", dcf.terminal.terminal_roc / 100, NF_PCT),
        ("RR = g/ROC", dcf.terminal.terminal_reinvestment_rate / 100, NF_PCT),
    ]
    for i, (label, val, nf) in enumerate(tv_items):
        r = 26 + i
        _cell(ws, r, 1, label, font=FONT_10, fill=FILL_TERMINAL, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT, fill=FILL_TERMINAL, border=BORDER_THIN, nf=nf)

    # Terminal Value = FCFF / (WACC - g)  (FORMUL)
    _cell(ws, 31, 1, "Terminal Value (TV)", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, 31, 2, f"=B26/(B27-B28)", font=FONT_FORMULA, fill=FILL_OUTPUT,
          border=BORDER_THIN, nf=nf_c,
          comment_text="TV = FCFF_terminal / (WACC - g)")

    # PV(TV)
    last_df_col = _col_letter(2 + n_years)
    _cell(ws, 32, 1, "PV(Terminal Value)", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, 32, 2, f"=B31*{last_df_col}22", font=FONT_FORMULA, fill=FILL_OUTPUT,
          border=BORDER_THIN, nf=nf_c,
          comment_text="PV(TV) = TV x son yil iskonto faktoru")

    # === ROW 34: DEGERLEME OZETI ===
    _merge_header(ws, 34, 1, 6, "DEGERLEME OZETi")

    # PV(10Y FCFF) = SUM formulu
    first_pv = _col_letter(3)
    last_pv = _col_letter(2 + n_years)
    _cell(ws, 35, 1, "PV (10Y FCFF)", font=FONT_10, border=BORDER_THIN)
    _cell(ws, 35, 2, f"=SUM({first_pv}23:{last_pv}23)", font=FONT_FORMULA,
          border=BORDER_THIN, nf=nf_c)

    _cell(ws, 36, 1, "PV (Terminal Value)", font=FONT_10, border=BORDER_THIN)
    _cell(ws, 36, 2, f"=B32", font=FONT_LINK, border=BORDER_THIN, nf=nf_c)

    _cell(ws, 37, 1, "= Faaliyet Varlik Degeri", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, 37, 2, f"=B35+B36", font=FONT_FORMULA, fill=FILL_OUTPUT, border=BORDER_THIN, nf=nf_c)

    # Equity Bridge (formuller)
    bridge_data = [
        ("(-) Borc", dcf.bridge.minus_debt, True),
        ("(-) Azinlik Paylari", dcf.bridge.minus_minority, True),
        ("(+) Nakit", dcf.bridge.plus_cash, True),
        ("(+) Istirakler", dcf.bridge.plus_cross_holdings, True),
        ("(-) Opsiyonlar", dcf.bridge.minus_options, True),
    ]
    for i, (label, val, is_input) in enumerate(bridge_data):
        r = 38 + i
        _cell(ws, r, 1, label, font=FONT_10, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT if is_input else FONT_FORMULA,
              fill=FILL_INPUT if is_input else FILL_WHITE, border=BORDER_THIN, nf=nf_c)

    # = Ozsermaye = Faaliyet - Borc - Azinlik + Nakit + Istirak - Opsiyon
    _cell(ws, 43, 1, "= OZSERMAYE DEGERi", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
    _cell(ws, 43, 2, f"=B37-B38-B39+B40+B41-B42", font=FONT_FORMULA,
          fill=FILL_OUTPUT, border=BORDER_THIN, nf=nf_c)

    # Pay sayisi
    _cell(ws, 44, 1, "Pay Sayisi (M)", font=FONT_10, border=BORDER_THIN)
    _cell(ws, 44, 2, dcf.bridge.shares, font=FONT_INPUT, fill=FILL_INPUT,
          border=BORDER_THIN, nf=NF_NUM1)

    # Hisse degeri (FORMUL)
    _cell(ws, 45, 1, "= HiSSE BASINA DEGER", font=FONT_OUTPUT, fill=FILL_OUTPUT,
          border=BORDER_MEDIUM)
    _cell(ws, 45, 2, f"=B43/B44", font=FONT_OUTPUT, fill=FILL_OUTPUT,
          border=BORDER_MEDIUM, nf=nf_c2)

    # Mevcut fiyat + potansiyel
    if params.get("current_price"):
        _cell(ws, 47, 1, "Mevcut Fiyat", font=FONT_10, border=BORDER_THIN)
        _cell(ws, 47, 2, params["current_price"], font=FONT_INPUT, fill=FILL_INPUT,
              border=BORDER_THIN, nf=nf_c2)
        _cell(ws, 48, 1, "Potansiyel", font=FONT_OUTPUT, fill=FILL_OUTPUT, border=BORDER_THIN)
        _cell(ws, 48, 2, f"=B45/B47-1", font=FONT_FORMULA, fill=FILL_OUTPUT,
              border=BORDER_THIN, nf=NF_PCT)

    # TV/EV check
    _cell(ws, 50, 1, "TV / EV", font=FONT_10, fill=FILL_WARN, border=BORDER_THIN)
    _cell(ws, 50, 2, f"=B36/B37", font=FONT_FORMULA, fill=FILL_WARN,
          border=BORDER_THIN, nf=NF_PCT)

    _apply_border_box(ws, 9, 1, 23, term_col)
    _apply_border_box(ws, 25, 1, 32, 6)
    _apply_border_box(ws, 34, 1, 50, 2)

    return ws


# ============================================================
# SHEET 3: WACC
# ============================================================

def _build_wacc(wb, wacc, params):
    """WACC hesaplama sheet'i - tum bilesenler."""
    ws = wb.create_sheet("WACC")
    cur = params.get("currency", "TL")

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 20

    _merge_header(ws, 1, 1, 3, "WACC HESAPLAMA - Damodaran Metodolojisi")

    # Cost of Equity
    _merge_header(ws, 3, 1, 3, "OZSERMAYE MALiYETi (Ke)")
    ke_items = [
        ("ABD 10Y Treasury (Ham)", wacc.rf_raw / 100, NF_PCT2, "ABD 10Y"),
        ("(-) Aa1 Default Spread", wacc.rf_default_adj / 100, NF_PCT2, "Damodaran"),
        ("= Risksiz Oran (Rf)", wacc.rf / 100, NF_PCT2, None),
        ("Kaldiracsiz Beta (B_U)", wacc.beta_u, NF_NUM2, "Damodaran sektor betasi"),
        ("D/E Orani", wacc.debt_equity, NF_NUM2, "Piyasa degeri bazli"),
        ("Kaldıracli Beta (B_L)", wacc.beta_l, NF_NUM2, None),
        ("Olgun Piyasa ERP", wacc.erp_mature / 100, NF_PCT2, "Damodaran"),
        ("Ulke Risk Primi (CRP)", wacc.crp / 100, NF_PCT2, "Blended: rating + CDS"),
        ("Lambda", wacc.lambda_factor, NF_NUM2, "Gelir dagilim faktoru"),
        ("= Ke (USD)", wacc.ke_usd / 100, NF_PCT2, None),
    ]
    for i, (label, val, nf, src) in enumerate(ke_items):
        r = 4 + i
        is_calc = src is None
        _cell(ws, r, 1, label, font=FONT_10B if is_calc else FONT_10,
              fill=FILL_OUTPUT if is_calc else FILL_WHITE, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_FORMULA if is_calc else FONT_INPUT,
              fill=FILL_OUTPUT if is_calc else FILL_INPUT, border=BORDER_THIN, nf=nf,
              comment_text=f"Kaynak: {src}" if src else None)

    # Cost of Debt
    _merge_header(ws, 15, 1, 3, "BORC MALiYETi (Kd)")
    kd_items = [
        ("Kd Kaynagi", wacc.kd_source, None, None),
        ("Kd Pre-tax (USD)", wacc.kd_pretax_usd / 100, NF_PCT2, None),
        ("Efektif Vergi Orani", wacc.tax_rate / 100, NF_PCT2, "KAP"),
        ("Kd After-tax (USD)", wacc.kd_aftertax_usd / 100, NF_PCT2, None),
    ]
    if wacc.synthetic_rating:
        kd_items.insert(1, ("Sentetik Rating", wacc.synthetic_rating, None, f"ICR: {wacc.synthetic_icr:.2f}"))

    for i, (label, val, nf, src) in enumerate(kd_items):
        r = 16 + i
        _cell(ws, r, 1, label, font=FONT_10, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN,
              nf=nf, comment_text=f"Kaynak: {src}" if src else None)

    # Agirliklar
    w_start = 16 + len(kd_items) + 1
    _merge_header(ws, w_start, 1, 3, "SERMAYE YAPISI")
    w_items = [
        ("E/V (Ozsermaye Agirligi)", wacc.equity_weight / 100, NF_PCT),
        ("D/V (Borc Agirligi)", wacc.debt_weight / 100, NF_PCT),
    ]
    for i, (label, val, nf) in enumerate(w_items):
        r = w_start + 1 + i
        _cell(ws, r, 1, label, font=FONT_10, border=BORDER_THIN)
        _cell(ws, r, 2, val, font=FONT_INPUT, fill=FILL_INPUT, border=BORDER_THIN, nf=nf)

    # WACC
    wacc_row = w_start + 4
    _merge_header(ws, wacc_row, 1, 3, "WACC SONUCU")
    results = [
        ("WACC (USD)", wacc.wacc_usd / 100, NF_PCT2),
    ]
    if wacc.wacc_tl is not None:
        results.extend([
            ("", None, None),
            ("Yurt ici Enflasyon Beklentisi", wacc.inflation_domestic / 100, NF_PCT),
            ("ABD Enflasyon Beklentisi", wacc.inflation_us / 100, NF_PCT),
            ("WACC (TL) - Fisher", wacc.wacc_tl / 100, NF_PCT2),
            ("Ke (TL)", wacc.ke_tl / 100, NF_PCT2),
            ("Kd Pre-tax (TL)", wacc.kd_pretax_tl / 100, NF_PCT2),
        ])
    if wacc.terminal_wacc_usd is not None:
        results.extend([
            ("", None, None),
            ("Terminal WACC (USD)", wacc.terminal_wacc_usd / 100, NF_PCT2),
        ])
        if wacc.terminal_wacc_tl is not None:
            results.append(("Terminal WACC (TL)", wacc.terminal_wacc_tl / 100, NF_PCT2))

    for i, (label, val, nf) in enumerate(results):
        r = wacc_row + 1 + i
        is_key = "WACC" in label and "Terminal" not in label and val is not None
        fill = FILL_OUTPUT if is_key else FILL_WHITE
        font = FONT_OUTPUT if is_key else FONT_10
        _cell(ws, r, 1, label, font=font, fill=fill, border=BORDER_THIN if val else None)
        if val is not None:
            _cell(ws, r, 2, val, font=font, fill=fill, border=BORDER_THIN, nf=nf)

    # CRP Detay
    if wacc.crp_rating_based is not None:
        crp_row = wacc_row + len(results) + 3
        _merge_header(ws, crp_row, 1, 3, "CRP DETAY")
        crp_items = [
            ("CRP (Rating Bazli)", wacc.crp_rating_based / 100, NF_PCT2),
        ]
        if wacc.crp_cds_based is not None:
            crp_items.append(("CRP (CDS Bazli)", wacc.crp_cds_based / 100, NF_PCT2))
        if wacc.crp_blend_weights:
            crp_items.append(("Agirlik: Rating / CDS",
                              f"{wacc.crp_blend_weights[0]:.0%} / {wacc.crp_blend_weights[1]:.0%}", None))
        crp_items.append(("CRP (Blended)", wacc.crp / 100, NF_PCT2))

        for i, (label, val, nf) in enumerate(crp_items):
            r = crp_row + 1 + i
            _cell(ws, r, 1, label, font=FONT_10, border=BORDER_THIN)
            _cell(ws, r, 2, val, font=FONT_10, fill=FILL_INPUT, border=BORDER_THIN, nf=nf)

    return ws


# ============================================================
# SHEET 4: SENSITIVITY
# ============================================================

def _build_sensitivity(wb, dcf, wacc_result, params):
    """3 sensitivity tablosu, 5x5, formul referansli."""
    ws = wb.create_sheet("SENSITIVITY")

    cur = params.get("currency", "TL")
    nf_c, nf_c2 = _nf_currency(cur)

    ws.column_dimensions['A'].width = 14
    for i in range(2, 9):
        ws.column_dimensions[_col_letter(i)].width = 14

    _merge_header(ws, 1, 1, 7, "DUYARLILIK ANALiZi (Sensitivity)")

    # Degerleri import et
    from dcf_engine import DCFEngine, sensitivity_grid

    # Engine parametrelerini yeniden olustur
    engine_params = params.get("engine_params", {})

    if not engine_params:
        # Fallback: DCFResult'tan geri hesapla
        _cell(ws, 3, 1, "Sensitivity grid icin engine_params gerekli.", font=FONT_WARN)
        _cell(ws, 4, 1, "generate_dcf_excel'e params['engine_params'] ekleyin.", font=FONT_SMALL)
        return ws

    # === TABLO 1: WACC vs Terminal Buyume ===
    base_wacc = params.get("wacc_for_sensitivity", wacc_result.wacc_tl or wacc_result.wacc_usd)
    base_tg = dcf.terminal.terminal_growth
    step_w = max(2.0, base_wacc * 0.05)  # %5 adim veya en az 2pp
    step_g = max(1.0, base_tg * 0.1)

    wacc_vals = [round(base_wacc - 2*step_w, 1), round(base_wacc - step_w, 1),
                 round(base_wacc, 1),
                 round(base_wacc + step_w, 1), round(base_wacc + 2*step_w, 1)]
    tg_vals = [round(base_tg - 2*step_g, 1), round(base_tg - step_g, 1),
               round(base_tg, 1),
               round(base_tg + step_g, 1), round(base_tg + 2*step_g, 1)]

    grid1 = sensitivity_grid(engine_params, "wacc", wacc_vals, "terminal_growth", tg_vals)

    r = 3
    _merge_header(ws, r, 1, 7, "WACC vs Terminal Buyume")
    r += 1
    _cell(ws, r, 1, "WACC \\ g", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
    for j, tg in enumerate(tg_vals):
        _cell(ws, r, 2 + j, f"{tg:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)

    for i, wv in enumerate(wacc_vals):
        r_data = r + 1 + i
        _cell(ws, r_data, 1, f"{wv:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)
        for j, tg in enumerate(tg_vals):
            val = grid1[i][j] if grid1[i][j] is not None else 0
            is_base = (abs(wv - base_wacc) < 0.05 and abs(tg - base_tg) < 0.05)
            fill = FILL_OUTPUT if is_base else FILL_WHITE
            font = FONT_OUTPUT if is_base else FONT_FORMULA
            _cell(ws, r_data, 2 + j, val, font=font, fill=fill,
                  border=BORDER_THIN, nf=nf_c2)

    _apply_border_box(ws, 3, 1, r + 6, 7)

    # === TABLO 2: Gelir Buyumesi Y1 vs EBIT Marj Hedefi ===
    base_g1 = dcf.projections[0].revenue_growth if dcf.projections else 10.0
    base_margin = dcf.terminal.terminal_ebit_margin

    g1_vals = [round(base_g1 - 6, 1), round(base_g1 - 3, 1), round(base_g1, 1),
               round(base_g1 + 3, 1), round(base_g1 + 6, 1)]
    margin_vals = [round(base_margin - 4, 1), round(base_margin - 2, 1),
                   round(base_margin, 1),
                   round(base_margin + 2, 1), round(base_margin + 4, 1)]

    # Buyume ve marj icin growth array'i adjust et
    grid2 = []
    for gv in g1_vals:
        row = []
        for mv in margin_vals:
            try:
                p = dict(engine_params)
                # Tum buyume oranlarini scale et
                if "revenue_growth" in p and p["revenue_growth"]:
                    base_first = p["revenue_growth"][0]
                    diff = gv - base_first
                    p["revenue_growth"] = [max(0, g + diff) for g in p["revenue_growth"]]
                elif "revenue_growth_y1_5" in p:
                    p["revenue_growth_y1_5"] = gv
                p["target_ebit_margin"] = mv
                eng = DCFEngine(**p)
                res = eng.calculate()
                row.append(round(res.bridge.value_per_share, 2))
            except Exception:
                row.append(None)
        grid2.append(row)

    r2 = r + 8
    _merge_header(ws, r2, 1, 7, "Y1 Gelir Buyumesi vs EBIT Marj Hedefi")
    r2 += 1
    _cell(ws, r2, 1, "Buyume \\ Marj", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
    for j, mv in enumerate(margin_vals):
        _cell(ws, r2, 2 + j, f"{mv:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)

    for i, gv in enumerate(g1_vals):
        rd = r2 + 1 + i
        _cell(ws, rd, 1, f"{gv:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)
        for j, mv in enumerate(margin_vals):
            val = grid2[i][j] if grid2[i][j] is not None else 0
            is_base = (abs(gv - base_g1) < 0.05 and abs(mv - base_margin) < 0.05)
            fill = FILL_OUTPUT if is_base else FILL_WHITE
            font = FONT_OUTPUT if is_base else FONT_FORMULA
            _cell(ws, rd, 2 + j, val, font=font, fill=fill,
                  border=BORDER_THIN, nf=nf_c2)

    _apply_border_box(ws, r2 - 1, 1, r2 + 6, 7)

    # === TABLO 3: Terminal WACC vs Terminal ROC ===
    tw = dcf.terminal.terminal_wacc
    troc = dcf.terminal.terminal_roc
    step_tw = max(1.0, tw * 0.05)
    step_roc = max(1.0, troc * 0.05)

    tw_vals = [round(tw - 2*step_tw, 1), round(tw - step_tw, 1), round(tw, 1),
               round(tw + step_tw, 1), round(tw + 2*step_tw, 1)]
    roc_vals = [round(troc - 2*step_roc, 1), round(troc - step_roc, 1),
                round(troc, 1),
                round(troc + step_roc, 1), round(troc + 2*step_roc, 1)]

    grid3 = []
    for twv in tw_vals:
        row = []
        for rv in roc_vals:
            try:
                p = dict(engine_params)
                p["terminal_wacc"] = twv
                p["terminal_roc"] = rv
                # wacc_schedule'u da guncelle (son deger = terminal)
                if "wacc_schedule" in p and p["wacc_schedule"]:
                    p["wacc_schedule"][-1] = twv
                eng = DCFEngine(**p)
                res = eng.calculate()
                row.append(round(res.bridge.value_per_share, 2))
            except Exception:
                row.append(None)
        grid3.append(row)

    r3 = r2 + 8
    _merge_header(ws, r3, 1, 7, "Terminal WACC vs Terminal ROC")
    r3 += 1
    _cell(ws, r3, 1, "WACC \\ ROC", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN)
    for j, rv in enumerate(roc_vals):
        _cell(ws, r3, 2 + j, f"{rv:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)

    for i, twv in enumerate(tw_vals):
        rd = r3 + 1 + i
        _cell(ws, rd, 1, f"{twv:.1f}%", font=FONT_10B, fill=FILL_SUB_HEADER,
              border=BORDER_THIN, alignment=ALIGN_C)
        for j, rv in enumerate(roc_vals):
            val = grid3[i][j] if grid3[i][j] is not None else 0
            is_base = (abs(twv - tw) < 0.05 and abs(rv - troc) < 0.05)
            fill = FILL_OUTPUT if is_base else FILL_WHITE
            font = FONT_OUTPUT if is_base else FONT_FORMULA
            _cell(ws, rd, 2 + j, val, font=font, fill=fill,
                  border=BORDER_THIN, nf=nf_c2)

    _apply_border_box(ws, r3 - 1, 1, r3 + 6, 7)

    return ws


# ============================================================
# SHEET 5: DIAGNOSTICS (17 Damodaran Sorusu)
# ============================================================

def _build_diagnostics(wb, dcf, wacc, params):
    """Damodaran 17 soru diagnostics sheet'i."""
    ws = wb.create_sheet("DIAGNOSTICS")

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 40

    _merge_header(ws, 1, 1, 3, "DAMODARAN DiAGNOSTiCS - 17 SORU")

    questions = [
        "1. Gelir buyume oraniniz sektor ortalamasindan yuksekse, sirketiniz kucuk mu?",
        "2. Gelir buyume tahmininiz, sirketin son yil buyumesinden cok farkliysa neden?",
        "3. Y10 geliriniz toplam pazarin ne kadari? Pazar bu kadar buyuk mu?",
        "4. En buyuk pazar oyuncularinin su anki geliri ne? Y10'da onlari asiyor musunuz?",
        "5. Pazardaki toplam buyume nedir? Buyumeden ne kadar pay aliyorsunuz?",
        "6. Sektorun EBIT marji nedir? Sizin marjiniz neden farkli?",
        "7. Isletmenin birim ekonomisi nedir? Ekstra 1 birim maliyeti ne kadar?",
        "8. Bu isteki rekabet nasil? Marji koruyabilecek misiniz?",
        "9. Buyumeniz toparlanma mi yoksa yeni buyume mi? S/C orani bunu yansitiyor mu?",
        "10. Fazla kapasite var mi? Yakin vadeli buyumeyi karsilamak icin ne kadar gerekli?",
        "11. Yatirim verimliligi sirket buyudukce degisiyor mu?",
        "12. Y10 sermaye getirisi (ROC) konusunda rahat misiniz?",
        "13. Yeniden yatiriminiz gelir buyume tahminiyle tutarli mi?",
        "14. Sermaye maliyetiniz sektor ortalamasiyla nasil karsilastiriliyor?",
        "15. Sermaye maliyeti zaman icinde neden degisiyor?",
        "16. Iflas oraniniz sirketin ozellikleriyle uyumlu mu?",
        "17. Hesapladiginiz deger piyasa fiyatiyla nasil karsilastiriliyor? Sapma kaynagi ne?",
    ]

    for i, q in enumerate(questions):
        r = 3 + i * 3  # Her soru icin 3 satir (soru + cevap + bosluk)
        _cell(ws, r, 1, f"{i+1}", font=FONT_10B, fill=FILL_SUB_HEADER, border=BORDER_THIN,
              alignment=ALIGN_C)
        _cell(ws, r, 2, q, font=FONT_10, border=BORDER_THIN, alignment=ALIGN_WRAP)
        _cell(ws, r, 3, "[Cevaplanacak]", font=FONT_SMALL, fill=FILL_INPUT,
              border=BORDER_THIN, alignment=ALIGN_WRAP)
        # Cevap satiri (genisletilmis)
        ws.merge_cells(start_row=r+1, start_column=2, end_row=r+1, end_column=3)
        _cell(ws, r+1, 2, "", font=FONT_10, fill=FILL_STORY, alignment=ALIGN_WRAP)
        ws.row_dimensions[r+1].height = 40

    return ws


# ============================================================
# ANA FONKSIYON
# ============================================================

def generate_dcf_excel(dcf_result, wacc_result, params: dict,
                       output_path: str) -> str:
    """
    DCF Excel dosyasi olustur.

    Args:
        dcf_result: DCFResult (dcf_engine.py'den)
        wacc_result: WACCResult (wacc_calculator.py'den)
        params: Ek parametreler dict:
            ticker: str
            company_name: str
            currency: str ("TL" veya "USD")
            date: str
            base_year: str
            current_price: float (opsiyonel)
            tax_rate: float (%)
            sales_to_capital: float
            growth_story: str
            profitability_story: str
            competitive_story: str
            engine_params: dict (DCFEngine constructor params, sensitivity icin)
            wacc_for_sensitivity: float (sensitivity grid'de kullanilacak WACC)
        output_path: Cikti dosyasi yolu

    Returns:
        Olusturulan dosyanin yolu
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Default sheet'i kaldir

    # Sheet'leri olustur
    _build_ozet(wb, dcf_result, wacc_result, params)
    _build_dcf(wb, dcf_result, wacc_result, params)
    _build_wacc(wb, wacc_result, params)
    _build_sensitivity(wb, dcf_result, wacc_result, params)
    _build_diagnostics(wb, dcf_result, wacc_result, params)

    # Kaydet
    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    print(f"DCF Excel olusturuldu: {output_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Boyut: {os.path.getsize(output_path) / 1024:.1f} KB")

    return output_path


# ============================================================
# CLI / DEMO
# ============================================================

def main():
    """EBEBK demo ile test."""
    if len(sys.argv) > 1 and sys.argv[1] == "--help":
        print("DCF Excel Generator v2.0")
        print("Kullanim: python dcf_excel_generator.py [--demo]")
        return

    from dcf_engine import DCFEngine
    from wacc_calculator import calculate_wacc

    # EBEBK verileri
    wacc_r = calculate_wacc(
        rf_10y_raw=4.50, beta_u=0.82, erp_mature=4.60,
        sovereign_rating="BB-", cds_spread_bps=280,
        lambda_factor=1.0, company_icr=2.81, tax_rate=23.0,
        market_cap=9896, total_debt=3385, cash=2854,
        inflation_domestic=25.0, inflation_us=2.5,
    )

    engine_params = dict(
        base_revenue=27675, base_ebit=2404, base_da=1136,
        revenue_growth=[43.9, 36.9, 28.9, 25.7, 21.3, 18.0, 14.8, 12.7, 11.6, 10.5],
        terminal_growth=10.5,
        target_ebit_margin=14.0, margin_convergence_year=7,
        sales_to_capital=5.65, sales_to_capital_terminal=4.0,
        tax_rate=23.0, terminal_tax_rate=25.0,
        wacc=wacc_r.wacc_tl, terminal_wacc=wacc_r.terminal_wacc_tl,
        terminal_roc=wacc_r.terminal_wacc_tl,
        total_debt=3385, cash=2854, minority_interest=0,
        shares=160, one_time_items=65,
        scenario_name="Base", currency="TL", base_year="FY2025",
        base_ebitda=3540, current_price=61.85,
    )

    engine = DCFEngine(**engine_params)
    dcf_result = engine.calculate()

    params = {
        "ticker": "EBEBK",
        "company_name": "Ebebek Magazacilik A.S.",
        "currency": "TL",
        "date": datetime.now().strftime("%d.%m.%Y"),
        "base_year": "FY2025",
        "current_price": 61.85,
        "tax_rate": 23.0,
        "sales_to_capital": 5.65,
        "growth_story": (
            "Turkiye'nin en buyuk bebek & anne urunleri perakendecisi. "
            "FY2025 hasilat 27.7B TL (+%51 YoY). 300 magaza + online (%49 buyume). "
            "2026 hedefi 330 magaza, 37B TL hasilat. "
            "UK operasyonu 3 magaza ile pilot asamada."
        ),
        "profitability_story": (
            "Sunum FAVOK marji %12.8. Vadeli alim finansman gideri 1.9B TL "
            "EBIT tanimini karistiriyor. Duzeltilmis EBIT marji ~%8.7. "
            "Olcek etkisi ve premiumlasma ile %14 hedef marjina yaklasim."
        ),
        "competitive_story": (
            "Moat skoru 3/10 - perakende, dusuk switching cost. "
            "Ozel markalarda (Bebebebek) brut marj avantaji. "
            "Online + fiziksel entegrasyon. "
            "Risk: CEO degisikligi (Halil Erdogmus ayrilisi)."
        ),
        "engine_params": engine_params,
        "wacc_for_sensitivity": wacc_r.wacc_tl,
    }

    output_path = os.path.expanduser(
        "~/.openclaw/workspace-kaya/research/companies/EBEBK/EBEBK_DCF_Model.xlsx"
    )
    generate_dcf_excel(dcf_result, wacc_r, params, output_path)
    print(f"\nHisse Degeri: {dcf_result.bridge.value_per_share:.2f} TL")
    print(f"Mevcut Fiyat: 61.85 TL")
    print(f"Potansiyel:   {(dcf_result.bridge.value_per_share / 61.85 - 1) * 100:.1f}%")


if __name__ == "__main__":
    main()
